from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import os, json, io, uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cantina-raiz-secret-2024")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get("DATABASE_URL", "sqlite:///cantina.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
# Chave Pix da cantina (configure no Railway em variáveis de ambiente)
PIX_CHAVE = os.environ.get("PIX_CHAVE", "cantina@colegioraiz.com.br")
PIX_NOME  = os.environ.get("PIX_NOME",  "Cantina Colegio Curso Raiz")

db = SQLAlchemy(app)

# ─── MODELOS ──────────────────────────────────────────────────────────────────

class Produto(db.Model):
    __tablename__ = "produtos"
    id         = db.Column(db.Integer, primary_key=True)
    nome       = db.Column(db.String(100), nullable=False)
    categoria  = db.Column(db.String(50), nullable=False)
    preco      = db.Column(db.Float, nullable=False)
    estoque    = db.Column(db.Integer, default=0)
    foto_b64   = db.Column(db.Text, nullable=True)
    ativo      = db.Column(db.Boolean, default=True)

class Pessoa(db.Model):
    __tablename__ = "pessoas"
    id           = db.Column(db.Integer, primary_key=True)
    nome         = db.Column(db.String(150), nullable=False)
    tipo         = db.Column(db.String(20), nullable=False)  # Aluno / Funcionario
    turma_cargo  = db.Column(db.String(100))
    contato      = db.Column(db.String(150))
    obs          = db.Column(db.Text)
    matricula    = db.Column(db.String(50), nullable=True)   # ex: MAT97431770123316
    vendas       = db.relationship("Venda", backref="pessoa", lazy=True)

class Venda(db.Model):
    __tablename__ = "vendas"
    id               = db.Column(db.Integer, primary_key=True)
    data             = db.Column(db.DateTime, default=datetime.now)
    total            = db.Column(db.Float, nullable=False)
    forma_pagamento  = db.Column(db.String(20), default="Dinheiro")
    pago             = db.Column(db.Boolean, default=True)
    pessoa_id        = db.Column(db.Integer, db.ForeignKey("pessoas.id"), nullable=True)
    itens            = db.relationship("ItemVenda", backref="venda", lazy=True, cascade="all, delete-orphan")

class ItemVenda(db.Model):
    __tablename__ = "itens_venda"
    id          = db.Column(db.Integer, primary_key=True)
    venda_id    = db.Column(db.Integer, db.ForeignKey("vendas.id"), nullable=False)
    produto_nome = db.Column(db.String(100))
    qtd         = db.Column(db.Integer)
    preco_unit  = db.Column(db.Float)
    subtotal    = db.Column(db.Float)

class SaldoAluno(db.Model):
    """Saldo de crédito de cada aluno na cantina."""
    __tablename__ = "saldo_aluno"
    id         = db.Column(db.Integer, primary_key=True)
    pessoa_id  = db.Column(db.Integer, db.ForeignKey("pessoas.id"), unique=True, nullable=False)
    saldo      = db.Column(db.Float, default=0.0)
    pessoa     = db.relationship("Pessoa", backref=db.backref("saldo_obj", uselist=False))

class RecargaSaldo(db.Model):
    """Registro de recargas (Pix) feitas pelos pais."""
    __tablename__ = "recargas_saldo"
    id          = db.Column(db.Integer, primary_key=True)
    pessoa_id   = db.Column(db.Integer, db.ForeignKey("pessoas.id"), nullable=False)
    valor       = db.Column(db.Float, nullable=False)
    data        = db.Column(db.DateTime, default=datetime.now)
    confirmado  = db.Column(db.Boolean, default=False)
    obs         = db.Column(db.String(200))
    pessoa      = db.relationship("Pessoa", backref="recargas")

class PedidoTablet(db.Model):
    """Pedido feito pelo aluno no tablet (autoatendimento)."""
    __tablename__ = "pedidos_tablet"
    id          = db.Column(db.Integer, primary_key=True)
    pessoa_id   = db.Column(db.Integer, db.ForeignKey("pessoas.id"), nullable=False)
    data        = db.Column(db.DateTime, default=datetime.now)
    total       = db.Column(db.Float, nullable=False)
    status      = db.Column(db.String(20), default="aguardando")  # aguardando / entregue
    itens_json  = db.Column(db.Text)   # JSON com lista de itens
    pessoa      = db.relationship("Pessoa", backref="pedidos_tablet")

class FechamentoCaixa(db.Model):
    __tablename__ = "fechamentos_caixa"
    id               = db.Column(db.Integer, primary_key=True)
    data             = db.Column(db.Date, default=date.today)
    hora_fechamento  = db.Column(db.String(10))
    total_dia        = db.Column(db.Float, default=0)
    total_dinheiro   = db.Column(db.Float, default=0)
    total_pix        = db.Column(db.Float, default=0)
    total_cartao     = db.Column(db.Float, default=0)
    total_apagar     = db.Column(db.Float, default=0)
    total_recebido   = db.Column(db.Float, default=0)
    qtd_vendas       = db.Column(db.Integer, default=0)
    observacoes      = db.Column(db.Text)

# ─── INICIALIZAÇÃO ─────────────────────────────────────────────────────────────

def seed_produtos():
    if Produto.query.count() == 0:
        produtos = [
            Produto(nome="Coxinha",          categoria="Salgado",  preco=4.50, estoque=20),
            Produto(nome="Pao de Queijo",    categoria="Salgado",  preco=3.00, estoque=30),
            Produto(nome="Esfiha",           categoria="Salgado",  preco=4.00, estoque=15),
            Produto(nome="Biscoito Recheado",categoria="Biscoito", preco=2.50, estoque=40),
            Produto(nome="Rosquinha",        categoria="Biscoito", preco=2.00, estoque=35),
            Produto(nome="Brigadeiro",       categoria="Doce",     preco=3.50, estoque=25),
            Produto(nome="Bolo Fatia",       categoria="Doce",     preco=5.00, estoque=12),
            Produto(nome="Suco de Laranja",  categoria="Bebida",   preco=5.00, estoque=20),
            Produto(nome="Agua",             categoria="Bebida",   preco=2.00, estoque=50),
            Produto(nome="Refrigerante",     categoria="Bebida",   preco=4.00, estoque=25),
        ]
        db.session.add_all(produtos)
        db.session.commit()

with app.app_context():
    db.create_all()
    seed_produtos()

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def borda_cel():
    s = Side(style='thin', color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def venda_to_dict(v):
    return {
        "id": v.id,
        "data": v.data.strftime("%d/%m/%Y %H:%M"),
        "total": v.total,
        "forma_pagamento": v.forma_pagamento,
        "pago": v.pago,
        "pessoa_id": v.pessoa_id,
        "pessoa_nome": v.pessoa.nome if v.pessoa else "",
        "itens": [{"nome": i.produto_nome, "qtd": i.qtd,
                   "preco_unit": i.preco_unit, "subtotal": i.subtotal} for i in v.itens]
    }

# ══════════════════════════════════════════════════════════════════════════════
# ROTAS PRINCIPAIS
# ══════════════════════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/cardapio")
def cardapio():
    return render_template("cardapio.html")

# ─── PRODUTOS ─────────────────────────────────────────────────────────────────

@app.route("/api/produtos")
def api_produtos():
    cat   = request.args.get("categoria", "Todos")
    busca = request.args.get("busca", "").lower()
    q = Produto.query.filter_by(ativo=True)
    if cat != "Todos":
        q = q.filter_by(categoria=cat)
    prods = q.all()
    if busca:
        prods = [p for p in prods if busca in p.nome.lower()]
    return jsonify([{
        "id": p.id, "nome": p.nome, "categoria": p.categoria,
        "preco": p.preco, "estoque": p.estoque, "foto_b64": p.foto_b64
    } for p in prods])

@app.route("/api/produtos/todos")
def api_produtos_todos():
    prods = Produto.query.filter_by(ativo=True).order_by(Produto.categoria, Produto.nome).all()
    return jsonify([{
        "id": p.id, "nome": p.nome, "categoria": p.categoria,
        "preco": p.preco, "estoque": p.estoque, "foto_b64": p.foto_b64
    } for p in prods])

@app.route("/api/produtos/salvar", methods=["POST"])
def salvar_produto():
    d = request.json
    if d.get("id"):
        p = Produto.query.get(d["id"])
        if not p: return jsonify({"erro": "Produto não encontrado"}), 404
    else:
        p = Produto()
        db.session.add(p)
    p.nome      = d["nome"]
    p.categoria = d["categoria"]
    p.preco     = float(d["preco"])
    p.estoque   = int(d["estoque"])
    if d.get("foto_b64"):
        p.foto_b64 = d["foto_b64"]
    db.session.commit()
    return jsonify({"ok": True, "id": p.id})

@app.route("/api/produtos/excluir/<int:pid>", methods=["DELETE"])
def excluir_produto(pid):
    p = Produto.query.get(pid)
    if p:
        p.ativo = False
        db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/produtos/estoque/<int:pid>", methods=["POST"])
def atualizar_estoque(pid):
    p = Produto.query.get(pid)
    if not p: return jsonify({"erro": "Não encontrado"}), 404
    p.estoque = int(request.json["estoque"])
    db.session.commit()
    return jsonify({"ok": True})

# ─── VENDAS ───────────────────────────────────────────────────────────────────

@app.route("/api/vendas/finalizar", methods=["POST"])
def finalizar_venda():
    d = request.json
    carrinho      = d["carrinho"]       # [{id, qtd}, ...]
    forma         = d["forma_pagamento"]
    pessoa_id     = d.get("pessoa_id")

    if forma == "A Pagar" and not pessoa_id:
        return jsonify({"erro": "Conta obrigatória para 'A Pagar'"}), 400

    total = 0.0
    itens_obj = []
    for item in carrinho:
        prod = Produto.query.get(item["id"])
        if not prod or prod.estoque < item["qtd"]:
            return jsonify({"erro": f"Estoque insuficiente: {prod.nome if prod else item['id']}"}), 400
        sub = prod.preco * item["qtd"]
        total += sub
        prod.estoque -= item["qtd"]
        itens_obj.append(ItemVenda(
            produto_nome=prod.nome, qtd=item["qtd"],
            preco_unit=prod.preco, subtotal=sub))

    venda = Venda(
        total=total, forma_pagamento=forma,
        pago=(forma != "A Pagar"),
        pessoa_id=pessoa_id if pessoa_id else None)
    db.session.add(venda)
    db.session.flush()
    for i in itens_obj:
        i.venda_id = venda.id
        db.session.add(i)
    db.session.commit()
    return jsonify({"ok": True, "total": total, "venda_id": venda.id})

@app.route("/api/vendas/historico")
def historico_vendas():
    ini = request.args.get("ini")  # YYYY-MM-DD
    fim = request.args.get("fim")
    q = Venda.query
    if ini:
        try:
            q = q.filter(Venda.data >= datetime.strptime(ini, "%Y-%m-%d"))
        except ValueError:
            pass
    if fim:
        try:
            from datetime import timedelta
            fim_dt = datetime.strptime(fim, "%Y-%m-%d") + timedelta(days=1)
            q = q.filter(Venda.data < fim_dt)
        except ValueError:
            pass
    vendas = q.order_by(Venda.data.desc()).limit(300).all()
    return jsonify([venda_to_dict(v) for v in vendas])

@app.route("/api/vendas/<int:vid>/mudar_pagamento", methods=["POST"])
def mudar_pagamento(vid):
    v = Venda.query.get(vid)
    if not v: return jsonify({"erro": "Venda não encontrada"}), 404
    nova_forma = request.json.get("forma", "").strip()
    formas_validas = ["Dinheiro", "Pix", "Cartão", "A Pagar", "Saldo"]
    if nova_forma not in formas_validas:
        return jsonify({"erro": f"Forma inválida. Use: {', '.join(formas_validas)}"}), 400
    v.forma_pagamento = nova_forma
    v.pago = (nova_forma != "A Pagar")
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/vendas/<int:vid>/excluir", methods=["DELETE"])
def excluir_venda(vid):
    v = Venda.query.get(vid)
    if not v: return jsonify({"erro": "Venda não encontrada"}), 404
    # Devolver estoque dos itens
    for item in v.itens:
        prod = Produto.query.filter_by(nome=item.produto_nome, ativo=True).first()
        if prod:
            prod.estoque += item.qtd
    # Se foi pago com Saldo, devolver crédito ao aluno
    if v.forma_pagamento == "Saldo" and v.pessoa_id:
        s = SaldoAluno.query.filter_by(pessoa_id=v.pessoa_id).first()
        if s:
            s.saldo += v.total
    db.session.delete(v)
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/vendas/<int:vid>/detalhes")
def detalhes_venda(vid):
    v = Venda.query.get(vid)
    if not v: return jsonify({"erro": "Venda não encontrada"}), 404
    return jsonify(venda_to_dict(v))

@app.route("/api/vendas/<int:vid>/corrigir", methods=["POST"])
def corrigir_venda(vid):
    """Substitui os itens de uma venda, ajustando o estoque corretamente."""
    v = Venda.query.get(vid)
    if not v:
        return jsonify({"erro": "Venda não encontrada"}), 404

    novos_itens = request.json.get("itens", [])  # [{id, qtd}, ...]
    if not novos_itens:
        return jsonify({"erro": "Nenhum item informado"}), 400

    # 1. Devolver estoque dos itens antigos
    for item_antigo in v.itens:
        prod = Produto.query.filter_by(nome=item_antigo.produto_nome, ativo=True).first()
        if prod:
            prod.estoque += item_antigo.qtd

    # 2. Remover itens antigos
    for item_antigo in list(v.itens):
        db.session.delete(item_antigo)
    db.session.flush()

    # 3. Registrar novos itens e descontar estoque
    total = 0.0
    for ni in novos_itens:
        prod = Produto.query.get(ni["id"])
        if not prod or not prod.ativo:
            db.session.rollback()
            return jsonify({"erro": f"Produto não encontrado: {ni['id']}"}), 400
        if prod.estoque < ni["qtd"]:
            db.session.rollback()
            return jsonify({"erro": f"Estoque insuficiente para: {prod.nome} (disponível: {prod.estoque})"}), 400
        sub = prod.preco * ni["qtd"]
        total += sub
        prod.estoque -= ni["qtd"]
        db.session.add(ItemVenda(
            venda_id=v.id,
            produto_nome=prod.nome,
            qtd=ni["qtd"],
            preco_unit=prod.preco,
            subtotal=sub
        ))

    # 4. Atualizar total da venda
    v.total = total
    db.session.commit()
    return jsonify({"ok": True, "total": total})

@app.route("/api/vendas/trocar_item", methods=["POST"])
def trocar_item():
    """Troca um item no carrinho (lógica no frontend, esta rota valida estoque)"""
    d = request.json
    prod_novo = Produto.query.get(d["produto_novo_id"])
    if not prod_novo:
        return jsonify({"erro": "Produto não encontrado"}), 404
    if prod_novo.estoque < d["qtd"]:
        return jsonify({"erro": f"Estoque insuficiente. Disponível: {prod_novo.estoque}"}), 400
    return jsonify({"ok": True, "preco": prod_novo.preco, "nome": prod_novo.nome})

# ─── PESSOAS / CONTAS ─────────────────────────────────────────────────────────

@app.route("/api/pessoas")
def api_pessoas():
    busca = request.args.get("busca","").lower()
    pessoas = Pessoa.query.order_by(Pessoa.nome).all()
    result = []
    for p in pessoas:
        if busca and busca not in p.nome.lower() and busca not in (p.turma_cargo or "").lower():
            continue
        devedor = db.session.query(db.func.sum(Venda.total)).filter(
            Venda.pessoa_id==p.id, Venda.pago==False).scalar() or 0
        result.append({
            "id": p.id, "nome": p.nome, "tipo": p.tipo,
            "turma_cargo": p.turma_cargo or "", "contato": p.contato or "",
            "obs": p.obs or "", "saldo_devedor": devedor,
            "matricula": p.matricula or ""
        })
    return jsonify(result)

@app.route("/api/pessoas/salvar", methods=["POST"])
def salvar_pessoa():
    d = request.json
    if d.get("id"):
        p = Pessoa.query.get(d["id"])
        if not p: return jsonify({"erro": "Não encontrado"}), 404
    else:
        p = Pessoa(); db.session.add(p)
    p.nome        = d["nome"]
    p.tipo        = d["tipo"]
    p.turma_cargo = d.get("turma_cargo","")
    p.contato     = d.get("contato","")
    p.obs         = d.get("obs","")
    p.matricula   = d.get("matricula") or None
    db.session.commit()
    return jsonify({"ok": True, "id": p.id})

@app.route("/api/pessoas/importar", methods=["POST"])
def importar_pessoas():
    """Importa alunos/funcionários da planilha TURMAS_RAIZ (aba MATRICULAS)."""
    if "arquivo" not in request.files:
        return jsonify({"erro": "Nenhum arquivo enviado"}), 400

    arquivo = request.files["arquivo"]
    try:
        wb = openpyxl.load_workbook(arquivo, data_only=True)
    except Exception:
        return jsonify({"erro": "Arquivo inválido. Envie um .xlsx"}), 400

    # Suporta tanto a planilha original (aba MATRICULAS) quanto planilha simples
    if "MATRICULAS" in wb.sheetnames:
        ws = wb["MATRICULAS"]
        # Cabeçalho na linha 2, dados a partir da linha 3
        # Col 4=NOME, 8=TURMA, 9=TURNO, 20=TELEFONE
        def extrair(row):
            nome     = row[3]
            turma    = row[7]
            turno    = row[8]
            tel      = row[19]
            if not nome or not str(nome).strip():
                return None
            tc = f"{turma} - {turno}" if turma and turno else (str(turma) if turma else "")
            return {
                "nome":        str(nome).strip(),
                "tipo":        "Aluno",
                "turma_cargo": tc.strip(),
                "contato":     str(tel).strip() if tel else "",
                "obs":         ""
            }
        linhas = list(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True))
    else:
        # Planilha simples: Nome | Tipo | Turma/Cargo | Contato
        ws = wb.active
        def extrair(row):
            if not row[0] or not str(row[0]).strip():
                return None
            return {
                "nome":        str(row[0]).strip(),
                "tipo":        str(row[1]).strip() if row[1] else "Aluno",
                "turma_cargo": str(row[2]).strip() if row[2] else "",
                "contato":     str(row[3]).strip() if row[3] else "",
                "obs":         ""
            }
        linhas = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

    inseridos = 0
    ignorados = 0
    nomes_existentes = {p.nome.lower() for p in Pessoa.query.all()}

    for row in linhas:
        dados = extrair(row)
        if not dados:
            continue
        if dados["nome"].lower() in nomes_existentes:
            ignorados += 1
            continue
        p = Pessoa(
            nome        = dados["nome"],
            tipo        = dados["tipo"],
            turma_cargo = dados["turma_cargo"],
            contato     = dados["contato"],
            obs         = dados["obs"]
        )
        db.session.add(p)
        nomes_existentes.add(dados["nome"].lower())
        inseridos += 1

    db.session.commit()
    return jsonify({"ok": True, "inseridos": inseridos, "ignorados": ignorados})


@app.route("/api/pessoas/modelo_excel")
def modelo_excel():
    """Gera planilha modelo para importação."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alunos"
    LA = "F47920"; RX = "7C4DFF"
    cabecalhos = ["Nome Completo", "Tipo (Aluno/Funcionario)", "Turma / Cargo", "Contato (Telefone)"]
    for c, h in enumerate(cabecalhos, 1):
        cel = ws.cell(row=1, column=c, value=h)
        cel.font = openpyxl.styles.Font(bold=True, color="FFFFFF", name="Calibri")
        cel.fill = openpyxl.styles.PatternFill("solid", fgColor=RX)
        cel.alignment = openpyxl.styles.Alignment(horizontal="center")
    # Exemplos
    exemplos = [
        ["João da Silva", "Aluno", "3º ANO - MANHÃ", "21 99999-9999"],
        ["Maria Souza",   "Aluno", "2º ANO - TARDE", "21 98888-8888"],
        ["Ana Lima",      "Funcionario", "Coordenação", "21 97777-7777"],
    ]
    for i, ex in enumerate(exemplos, 2):
        for c, v in enumerate(ex, 1):
            ws.cell(row=i, column=c, value=v).font = openpyxl.styles.Font(name="Calibri")
    for i, w in enumerate([30, 22, 22, 20], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="modelo_importacao_alunos.xlsx")


@app.route("/api/pessoas/excluir/<int:pid>", methods=["DELETE"])
def excluir_pessoa(pid):
    p = Pessoa.query.get(pid)
    if p: db.session.delete(p)
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/pessoas/<int:pid>/extrato")
def extrato_pessoa(pid):
    p = Pessoa.query.get(pid)
    if not p: return jsonify({"erro": "Não encontrado"}), 404
    vendas = Venda.query.filter_by(pessoa_id=pid).order_by(Venda.data.desc()).all()
    return jsonify({
        "pessoa": {"id": p.id, "nome": p.nome, "tipo": p.tipo,
                   "turma_cargo": p.turma_cargo or "", "contato": p.contato or ""},
        "vendas": [venda_to_dict(v) for v in vendas]
    })

@app.route("/api/pessoas/<int:pid>/marcar_pago/<int:vid>", methods=["POST"])
def marcar_pago(pid, vid):
    v = Venda.query.get(vid)
    if not v: return jsonify({"erro": "Venda não encontrada"}), 404
    v.pago = True
    v.forma_pagamento = request.json.get("forma","Dinheiro")
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/pessoas/<int:pid>/quitar_tudo", methods=["POST"])
def quitar_tudo(pid):
    forma = request.json.get("forma","Dinheiro")
    vendas = Venda.query.filter_by(pessoa_id=pid, pago=False).all()
    for v in vendas:
        v.pago = True
        v.forma_pagamento = forma
    db.session.commit()
    return jsonify({"ok": True, "qtd": len(vendas)})

# ─── FECHAMENTO DE CAIXA ──────────────────────────────────────────────────────

@app.route("/api/caixa/hoje")
def caixa_hoje():
    hoje = date.today()
    inicio = datetime(hoje.year, hoje.month, hoje.day, 0, 0, 0)
    fim    = datetime(hoje.year, hoje.month, hoje.day, 23, 59, 59)
    vendas = Venda.query.filter(Venda.data >= inicio, Venda.data <= fim).all()
    totais = {"Dinheiro":0.0,"Pix":0.0,"Cartão":0.0,"A Pagar":0.0}
    qtd    = {"Dinheiro":0,"Pix":0,"Cartão":0,"A Pagar":0}
    for v in vendas:
        f = v.forma_pagamento or "Dinheiro"
        if f in totais:
            totais[f] += v.total
            qtd[f] += 1
    total_recebido = totais["Dinheiro"] + totais["Pix"] + totais["Cartão"]
    ja_fechado = FechamentoCaixa.query.filter_by(data=hoje).first() is not None
    return jsonify({
        "data": hoje.strftime("%d/%m/%Y"),
        "totais": totais, "qtd": qtd,
        "total_dia": sum(totais.values()),
        "total_recebido": total_recebido,
        "total_apagar": totais["A Pagar"],
        "qtd_vendas": len(vendas),
        "ja_fechado": ja_fechado,
        "vendas": [venda_to_dict(v) for v in vendas]
    })

@app.route("/api/caixa/fechar", methods=["POST"])
def fechar_caixa():
    d = request.json
    hoje = date.today()
    f = FechamentoCaixa(
        data=hoje,
        hora_fechamento=datetime.now().strftime("%H:%M:%S"),
        total_dia=d["total_dia"],
        total_dinheiro=d["totais"]["Dinheiro"],
        total_pix=d["totais"]["Pix"],
        total_cartao=d["totais"]["Cartão"],
        total_apagar=d["totais"]["A Pagar"],
        total_recebido=d["total_recebido"],
        qtd_vendas=d["qtd_vendas"],
        observacoes=d.get("observacoes","")
    )
    db.session.add(f)
    db.session.commit()
    return jsonify({"ok": True})

@app.route("/api/caixa/historico")
def historico_caixa():
    mes = request.args.get("mes", 0, type=int)
    ano = request.args.get("ano", 0, type=int)
    q = FechamentoCaixa.query
    if mes > 0: q = q.filter(db.extract("month", FechamentoCaixa.data) == mes)
    if ano > 0: q = q.filter(db.extract("year",  FechamentoCaixa.data) == ano)
    fechamentos = q.order_by(FechamentoCaixa.data.desc()).all()
    return jsonify([{
        "id": f.id,
        "data": f.data.strftime("%d/%m/%Y"),
        "hora_fechamento": f.hora_fechamento,
        "total_dia":      f.total_dia,
        "total_dinheiro": f.total_dinheiro,
        "total_pix":      f.total_pix,
        "total_cartao":   f.total_cartao,
        "total_apagar":   f.total_apagar,
        "total_recebido": f.total_recebido,
        "qtd_vendas":     f.qtd_vendas,
        "observacoes":    f.observacoes or ""
    } for f in fechamentos])

@app.route("/api/caixa/anos")
def anos_caixa():
    anos = db.session.query(db.func.extract("year", FechamentoCaixa.data)).distinct().all()
    return jsonify(sorted([int(a[0]) for a in anos], reverse=True) or [date.today().year])

# ─── EXPORTAÇÕES EXCEL ────────────────────────────────────────────────────────

@app.route("/api/export/estoque")
def export_estoque():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Estoque"
    LA="F47920"; MA="8B4513"; bd=borda_cel()
    ws.merge_cells("A1:F1"); ws["A1"]="Colégio Curso Raiz - Relatório de Estoque"
    ws["A1"].font=Font(bold=True,size=16,color="FFFFFF",name="Calibri")
    ws["A1"].fill=PatternFill("solid",fgColor=LA)
    ws["A1"].alignment=Alignment(horizontal="center"); ws.row_dimensions[1].height=32
    ws.merge_cells("A2:F2")
    ws["A2"]="Gerado em: {}".format(datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    ws["A2"].font=Font(italic=True,size=10,name="Calibri"); ws["A2"].alignment=Alignment(horizontal="center")
    for c,h in enumerate(["Produto","Categoria","Preço (R$)","Estoque","Valor Total (R$)","Situação"],1):
        cel=ws.cell(row=4,column=c,value=h)
        cel.font=Font(bold=True,color="FFFFFF",name="Calibri",size=11)
        cel.fill=PatternFill("solid",fgColor=MA); cel.alignment=Alignment(horizontal="center")
    tg=0.0; linha=5
    for p in Produto.query.filter_by(ativo=True).order_by(Produto.categoria,Produto.nome).all():
        vt=p.preco*p.estoque; tg+=vt
        if p.estoque==0: sit,fc="Esgotado","FFEBEE"
        elif p.estoque<=5: sit,fc="Baixo","FFF9C4"
        else: sit,fc="OK","E8F5E9"
        for c,v in enumerate([p.nome,p.categoria,p.preco,p.estoque,vt,sit],1):
            cel=ws.cell(row=linha,column=c,value=v)
            cel.fill=PatternFill("solid",fgColor=fc); cel.border=bd
            cel.font=Font(name="Calibri",size=10); cel.alignment=Alignment(horizontal="center")
            if c in (3,5): cel.number_format='R$ #,##0.00'
        linha+=1
    ws.cell(row=linha+1,column=4,value="TOTAL:").font=Font(bold=True,name="Calibri")
    ct=ws.cell(row=linha+1,column=5,value=tg)
    ct.font=Font(bold=True,color=LA,size=12,name="Calibri"); ct.number_format='R$ #,##0.00'
    for i,l in enumerate([25,14,14,10,18,12],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=l
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="estoque_raiz.xlsx")

@app.route("/api/export/extrato/<int:pid>")
def export_extrato(pid):
    p = Pessoa.query.get(pid)
    if not p: return "Não encontrado", 404
    vendas_p = Venda.query.filter_by(pessoa_id=pid).order_by(Venda.data).all()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Extrato"
    LA="F47920"; RX="7C4DFF"; VD="5BAD6F"; bd=borda_cel()
    ws.merge_cells("A1:G1"); ws["A1"]="Colégio Curso Raiz - Extrato de Conta"
    ws["A1"].font=Font(bold=True,size=16,color="FFFFFF",name="Calibri")
    ws["A1"].fill=PatternFill("solid",fgColor=LA)
    ws["A1"].alignment=Alignment(horizontal="center"); ws.row_dimensions[1].height=32
    for i,(lbl,val) in enumerate([("Nome:",p.nome),("Tipo:",p.tipo),
        ("Turma/Cargo:",p.turma_cargo or "---"),("Contato:",p.contato or "---"),
        ("Gerado em:",datetime.now().strftime("%d/%m/%Y %H:%M"))],2):
        ws.cell(row=i,column=1,value=lbl).font=Font(bold=True,name="Calibri",color=RX)
        ws.cell(row=i,column=2,value=val).font=Font(name="Calibri")
    for c,h in enumerate(["Data/Hora","Item","Qtd","Preço Unit.","Subtotal","Forma Pag.","Status"],1):
        cel=ws.cell(row=8,column=c,value=h)
        cel.font=Font(bold=True,color="FFFFFF",name="Calibri",size=11)
        cel.fill=PatternFill("solid",fgColor=RX); cel.alignment=Alignment(horizontal="center")
    linha=9; tg=tp=td=0.0
    for v in vendas_p:
        pago=v.pago; fc="E8F5E9" if pago else "FFF3CD"
        st="Pago" if pago else "Pendente"; forma=v.forma_pagamento or "---"
        for item in v.itens:
            sub=item.subtotal; tg+=sub
            if pago: tp+=sub
            else: td+=sub
            for c,val in enumerate([v.data.strftime("%d/%m/%Y %H:%M"),item.produto_nome,
                item.qtd,item.preco_unit,sub,forma,st],1):
                cel=ws.cell(row=linha,column=c,value=val)
                cel.fill=PatternFill("solid",fgColor=fc); cel.border=bd
                cel.font=Font(name="Calibri",size=10); cel.alignment=Alignment(horizontal="center")
                if c in (4,5): cel.number_format='R$ #,##0.00'
            linha+=1
    for lbl,val,cor in [("Total Geral:",tg,LA),("Total Pago:",tp,VD),("Saldo Devedor:",td,"CC3300")]:
        ws.cell(row=linha+1,column=5,value=lbl).font=Font(bold=True,name="Calibri")
        c2=ws.cell(row=linha+1,column=6,value=val)
        c2.font=Font(bold=True,color=cor,size=12,name="Calibri"); c2.number_format='R$ #,##0.00'
        linha+=1
    for i,l in enumerate([22,20,6,12,12,14,12],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=l
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"extrato_{p.nome}.xlsx")

@app.route("/api/export/fechamento/<int:fid>")
def export_fechamento(fid):
    f = FechamentoCaixa.query.get(fid)
    if not f: return "Não encontrado", 404
    hoje_str = f.data.strftime("%d/%m/%Y")
    inicio = datetime(f.data.year, f.data.month, f.data.day)
    fim    = datetime(f.data.year, f.data.month, f.data.day, 23, 59, 59)
    vendas = Venda.query.filter(Venda.data >= inicio, Venda.data <= fim).all()
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Fechamento"
    LA="F47920"; VD="2E7D32"; VM="CC0000"; PI="32BCAD"; bd=borda_cel()
    ws.merge_cells("A1:F1"); ws["A1"]="COLÉGIO CURSO RAIZ - FECHAMENTO DE CAIXA"
    ws["A1"].font=Font(bold=True,size=16,color="FFFFFF",name="Calibri")
    ws["A1"].fill=PatternFill("solid",fgColor=LA)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=36
    ws.merge_cells("A2:F2")
    ws["A2"]="Data: {}  |  Fechado às: {}".format(hoje_str, f.hora_fechamento)
    ws["A2"].font=Font(italic=True,size=10,name="Calibri",color="888888")
    ws["A2"].alignment=Alignment(horizontal="center")
    cores_f={"Dinheiro":VD,"Pix":PI,"Cartão":"1565C0","A Pagar":"7C4DFF"}
    for c,h in enumerate(["Forma","Qtd Vendas","Total (R$)"],1):
        cel=ws.cell(row=4,column=c,value=h)
        cel.font=Font(bold=True,color="FFFFFF",name="Calibri")
        cel.fill=PatternFill("solid",fgColor="666666"); cel.alignment=Alignment(horizontal="center")
    for i,(forma,val,qtd) in enumerate([
        ("Dinheiro",f.total_dinheiro,0),("Pix",f.total_pix,0),
        ("Cartão",f.total_cartao,0),("A Pagar",f.total_apagar,0)],5):
        cor=cores_f.get(forma,"888888")
        ws.cell(row=i,column=1,value=forma).font=Font(bold=True,name="Calibri",color=cor)
        ws.cell(row=i,column=2,value=qtd).alignment=Alignment(horizontal="center")
        c3=ws.cell(row=i,column=3,value=val)
        c3.number_format='R$ #,##0.00'; c3.font=Font(bold=True,name="Calibri",color=cor)
        for c in range(1,4): ws.cell(row=i,column=c).border=bd
    for c,h in enumerate(["Horário","Itens","Total","Forma","Conta","Status"],1):
        cel=ws.cell(row=11,column=c,value=h)
        cel.font=Font(bold=True,color="FFFFFF",name="Calibri")
        cel.fill=PatternFill("solid",fgColor=LA); cel.alignment=Alignment(horizontal="center")
    linha=12
    for v in vendas:
        itens_s=", ".join("{} x{}".format(i.produto_nome,i.qtd) for i in v.itens)
        forma=v.forma_pagamento or "---"; status="Pago" if v.pago else "Pendente"
        conta=v.pessoa.nome if v.pessoa else "À vista"
        fc="E8F5E9" if v.pago else "FFF3CD"
        for c,val in enumerate([v.data.strftime("%H:%M"),itens_s,v.total,forma,conta,status],1):
            cel=ws.cell(row=linha,column=c,value=val)
            cel.fill=PatternFill("solid",fgColor=fc); cel.border=bd
            cel.font=Font(name="Calibri",size=10); cel.alignment=Alignment(horizontal="center")
            if c==3: cel.number_format='R$ #,##0.00'
        linha+=1
    for i,l in enumerate([10,40,14,14,20,10],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=l
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"fechamento_{hoje_str.replace('/','_')}.xlsx")

@app.route("/api/export/mensal")
def export_mensal():
    mes = request.args.get("mes", 0, type=int)
    ano = request.args.get("ano", date.today().year, type=int)
    q = FechamentoCaixa.query.filter(db.extract("year", FechamentoCaixa.data)==ano)
    if mes > 0:
        q = q.filter(db.extract("month", FechamentoCaixa.data)==mes)
    fechamentos = q.order_by(FechamentoCaixa.data).all()
    if not fechamentos:
        return "Sem dados", 404
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="Relatório Mensal"
    LA="F47920"; PI="32BCAD"; bd=borda_cel()
    ws.merge_cells("A1:G1"); ws["A1"]="COLÉGIO CURSO RAIZ - RELATÓRIO MENSAL DE CAIXA"
    ws["A1"].font=Font(bold=True,size=16,color="FFFFFF",name="Calibri")
    ws["A1"].fill=PatternFill("solid",fgColor=LA)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=36
    ws.merge_cells("A2:G2")
    ws["A2"]="Período: {}/{}  |  Gerado em: {}".format(
        mes if mes else "Todos", ano, datetime.now().strftime("%d/%m/%Y %H:%M"))
    ws["A2"].font=Font(italic=True,size=10,name="Calibri",color="888888")
    ws["A2"].alignment=Alignment(horizontal="center")
    for c,h in enumerate(["Data","Dinheiro","Pix","Cartão","A Pagar","Total Recebido","Total Dia"],1):
        cel=ws.cell(row=4,column=c,value=h)
        cel.font=Font(bold=True,color="FFFFFF",name="Calibri",size=11)
        cel.fill=PatternFill("solid",fgColor=PI); cel.alignment=Alignment(horizontal="center")
    t_d=t_p=t_c=t_a=t_r=t_t=0.0
    for i,f in enumerate(fechamentos,5):
        fc="E8F5E9" if i%2==0 else "FFFFFF"
        t_d+=f.total_dinheiro; t_p+=f.total_pix; t_c+=f.total_cartao
        t_a+=f.total_apagar; t_r+=f.total_recebido; t_t+=f.total_dia
        for c,v in enumerate([f.data.strftime("%d/%m/%Y"),f.total_dinheiro,f.total_pix,
            f.total_cartao,f.total_apagar,f.total_recebido,f.total_dia],1):
            cel=ws.cell(row=i,column=c,value=v)
            cel.fill=PatternFill("solid",fgColor=fc); cel.border=bd
            cel.font=Font(name="Calibri",size=10); cel.alignment=Alignment(horizontal="center")
            if c>1: cel.number_format='R$ #,##0.00'
    lt=len(fechamentos)+5+1
    for c,v in enumerate(["TOTAL",t_d,t_p,t_c,t_a,t_r,t_t],1):
        cel=ws.cell(row=lt,column=c,value=v)
        cel.font=Font(bold=True,color="FFFFFF",size=12,name="Calibri")
        cel.fill=PatternFill("solid",fgColor=LA); cel.border=bd
        cel.alignment=Alignment(horizontal="center")
        if c>1: cel.number_format='R$ #,##0.00'
    for i,l in enumerate([12,14,14,14,14,18,14],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=l
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"relatorio_mensal_{ano}.xlsx")


# ══════════════════════════════════════════════════════════════════════════════
# SISTEMA DE SALDO / QR CODE (identificação por matrícula do aluno)
# ══════════════════════════════════════════════════════════════════════════════

LOGO_B64_GLOBAL = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAMCAgICAgMCAgIDAwMDBAYEBAQEBAgGBgUGCQgKCgkICQkKDA8MCgsOCwkJDRENDg8QEBEQCgwSExIQEw8QEBD/2wBDAQMDAwQDBAgEBAgQCwkLEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD/wAARCAH0AfQDASIAAhEBAxEB/8QAHQABAAMAAgMBAAAAAAAAAAAAAAYHCAEFAgMECf/EAFsQAAEDAwEEBQQJDQ0FCQEBAAEAAgMEBREGBxIhMQgTQVFhInGBkRQyNjd1obGysxUXI0JSVnJzdHbB0dIWGCQ0NThTVFVik5S0M0OCouElJ0VGY2SEksJl0//EABwBAQABBQEBAAAAAAAAAAAAAAAHAQQFBggCA//EAEYRAAIBAwEEBgUJBQgCAgMAAAABAgMEEQUGEiExB0FRYXGRE4GhsdEUFyIyNlNyssEVMzVC0iM0UmKCkuHwFlQkJUOi8f/aAAwDAQACEQMRAD8AlaIi4nO0wiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIDjjlchpJ4LhaV2BbCad1NDrXWVHvySYfR0creDW9j3g8yeYCz+z+gXW0N2ra2Xe2+SXa/8AvE1/aDaC12etHc3L48klzb7F+r6iqtH7C9oWs4BW0NrbSUr+LZq1xia4d7RguI8cYU1/ek6y3M/ugtm/j2u6/HrWsGRsiaGMYGtHAALy5j/qpwtOi/RqNJRr705dbzjyS5EIXfSdrVao5UXGEepYz5t/8GFtabENoGiKd1bcra2qo2+2qKMmRrR3uGA5vnIwoCv0jlhinidFKxr2OBDmuGQQewrMW3nYGLcZtZaKpD7GyZKuhjHtO98Y7u8dnMLTNrejh6bSd5pbcoL60XxaXau1e03LZPpHWoVVZ6qlGT+rJcE32NdT7+RnpERRHglxPIREQqEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREKcjjzIM9qktg2ba61RCKix6ZramF3AShm4w/wDE7A+NW/sh6Ntzfc3XbaLazT01M4dVROka7rnc8uLSfJHd2rY9K2X1PV60KdGlJRl/M00ku1vBrurbU6XpFGdSrWi5R/lTTk32Yz//AApnTug9Yare1lh0/WVbXnHWNjIjHneeA9aklXsD2r0UBqJtKyFrRkhk0bnehrXElbcorfRW+BlNR0kcEMbd1jI2hrWjuAC+ktB4YUtW/RNYKlivWk59qwl5YfvIluOljUZVc0KMFDseW/PK9xkzYJsUrrxf333WFqnpqS2SDq6apiLDNNz8prh7Uc/OtYMaxg3GNwBwAC5DGgYAA8FycBb1s7s9a7OWvye34tvLk+bff4I0baHaG62iuvlNxwSWFFcku7x5ni9zI277yGtaMkk8AFELFtW0fqLUFTpy23AOqoHFrC4brJiPbbh7cfH2KtNvO1VwdNoewzkH2tfM094/2Q/T6lRNPUT0lRHVUszopYXB7HsOC1w5EFeb7XFb11SpLKXP4I0S81j0FZU6aylz+CN5+ZeL42SMcx7Q4EYIPJV7sg2lw68s5pq1zWXWiaG1DM/7QdkgHce3xVieCzlCtC5pqpDimZmhWjWgqlN8GZE28bF6+xalZdtJWqoqaC6vP2CCMvME3MgNbx3TzHpCh0Gw3anUQioZpOpDCM4e5rXeonK3Wd13AgH0JgAcgo6vejDTL27nc78oqTzuxwkn14yn1km2PSdqljawttyMnFY3pZba6uTXVwPzwvmk9Saak6u+2SsojnAMsTmtd5jyK6hfozd7Jar7QyW67UEFVTSjD45WAtPrWW9qHRp1DQXn2Vs+tz6+31BLjD1rGugd3ZcRlvctD2l6NrvS4ensG6sOtY+kvUuZvWzXSTa6pP0GoJUp9Tb+i/W+T8fMorPiuVJdQ7NddaVjM1+0xW00TeBl3d9g/wCJuR8ajSjm5tK9pLcuIOD7Gmn7SR7a8t7yG/bzU12ppr2BERWpdBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQHGE496KSaB0He9oV+jslmjx9vNM4eRDH2uP6B2q5tbWte1o0KEXKcnhJc2y2urqlZUZV68lGEVlt8kiNhpPABaO2AbCqOtoodaawoxKJfLo6SRvk7vY947fAK0ND7B9B6OpowbXFca1oBfU1TA8l391p4AKxo444WCONga1owABgAKctk+jf8AZ9aN5qjUmlwjzSfa31495B21nSO9RoystMTjFvjN8G12Jc1nt54PGKnggjEUMTWMaMNDQAAO4Be7hzXrklZEx0shAa0EknkAO1fNbLtbLzSMrrVXQVUD/ayQvDmn0hS0t2D3URM5ZfF8T7HEDivioLtbLp1v1Orqeo6h5ik6mQO3HjgWnHIhRXa5rA6N0bVVtPIWVlT/AAalI7JHA+V6GgnzgLLGm9W37Sl0F2s1dJHOXZlBOWzDOSHjtWKvtWhY1o02s559xir3U4WdWMGs9vcbeHHsUU2m6rbo3SFdeW4NRu9TTN75X8Gn0cT6F6dmu0O37QLL7Nhb1NXT4ZVU5PtHY5jvaexVj0n7q8vslljc4MIlqXgHgTwaMjw4+tfa8vIws5V6TzlcPXwPtd3cYWrrU3zXD1lFTTTVE0lRUSukllcXve45LnE5JJ78rxRd9c9O+wNIWbUZkJNznqItzd5CPHHPr9Sj+MJVE2uri/PH6mkxhKom11cWc6E1VUaM1RRX2Fz+rieG1DG/bxHg4eriPEBbQpKuCupIq2neHxTMbIxw5FrhkH1FYQWudiF2mu2ze1y1MhfJTtfTOJ7mOIaP/rurZdnLl70qD5c1+pn9CrtSlRfLmTzDTxHavjF1tbrgbS2vpzWNZ1hpxIOsDO/d54VV7ZNsR0uX6a05Kx10e37NMOPsYEcBj7sg58FQ2ndX3ew6rptV+ypZ6qObfmdI8l0zTwc0k94/Qsjd61StqyopZ48X2F/c6tSoVVSis8ePcbaOe1BjBXzW2vp7nQ09wpHh8NTG2aNw7WuGR8q9FXfLRQ11NbKu400VXWEtggfIA+QgZ8kcyszvRxnPAyu8sZzwPsmghqI3RTRNexww5rhkEeZZq2/7DKKho5da6Po+qEfl1tJG3ycH7do7PELTI4DhyXg+NkjSyRoc1wwQRkELCa9oFpr9pK2uFx6pY4p9qM5oOu3eg3Ubm3fDrjnhJdj/AEfUfm4QRwPBcchxW3ddbCNC6xpJOrtkNtriCW1NKwMO9/eaOBHxrIWudFXjQV+msV5jw5nlRytHkSsPJzfBc77S7GX+zbVSriVN8FJdvY11M6G2a20sdpM0qeYVFxcX2dqfWvadAiItNNzCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIDjt8y2h0c9FU+ltBU1yfDu1t4HsqYubhwafat8wHyrG1DC2orqenf7WWVjD5iQF+idnoY7bbKW3w46unhZGwdwDQB8il3on06Fe8rXk1lwSS7m88fYRF0sajOjaULODwptt96jjh5vPqPt5cuxeO+3O7kby8jwafBZG11r7U0Ov71WWTUFdTRtqXRRtjmIaAzyfa8uYPYprv9QhYRjKSzl4Oe729jZRUpLOWak1TQ1N005c7bRP3KippZYo3ZxhxaQFQ+xGyazs19o66Bkn1NqamaiuVMc5gcxhc1z2nlxAwfR2hejSHSNv1vljptV0rK+l4NdNENyZvj3O+JX5pvUdk1TbWXex1UdRBNzc0Yc0j7Vw5gjxVnTqW+qVYVac2pR6ustYToajUjVhJqUerrKe6UNRIKew0zd7ddJM89x4NCpGgsdwuVtuN0pI2ugtbI5Kjj5Qa926CB28ea0l0hNJV2otKw3G2wulntMjpXRtGS6Jww7A7SMA+bKqzYHAy56gu+nagE09ztc0UoPaMgD5xWH1O1dbUd2fKXLyx7zF6hbupfbsuUuXl8TqdjGop9P7QLc1shbT3GQUc7exwfwbw7w7d9ZUv6TkD26jtFRunq3Ub2g+IfyUa2daCvEu1OGyyQuxYq0TVkmCA1sbstP8AxENwO4lXntm0FJrfSuKBgdcrc4z0oJxv8PKZ6Ry8QF9LS2r1tOqUscnw9XNHu2t61WwqU8cnw9XNGSla+vLHPR7GdHTubjqpZHvx/wCrkt+JVXLT1EE7qWaCSOZjzG6NzSHNcDjBB5HK0xtgsBh2MwUbHE/UplK4nvDQGH5ysNPt3OjXfYv1z+hZ2NDfpVX2L9c/oZjPJaY2LVh09sfqb5UtJjifU1LAeGQ0YwPOWkLP+lNK3XWN5gstpicZJHDrJN3LYmdr3eAWuTo2ip9CyaKogGQmhdSsJH2xaRvHxLjlXmhW9VynXS5JpeJc6NQqNyrLqTS8THVdWXC/3aauqXOnra+cvdgZL3uPID0gAeZeV7s1bp+6VFnuQYKmmcGyBjt4AkA4z6Qp5sj0FcKzaP7Cu1I+H6hPM9TG8cntPkD0nBB8AotqSnrr9ry60tBDJU1FXc52QxsGS7MjsD1LF1LWfolUmnvSlhern7THTt5qCqS5t48uftNMbJ6uuk2UWqoiZ19RFSStiaTjeLHODG/EAqh05pjW9RtXstwvzZJKyZwuU/lEmnhO9gPz7XuDVf8AoywjS+lrbYfJLqOnYx5byL+bj6XErqNebTdM6AhLq5wnrpG5jpIsdY4dhd9y3xK26taxdCnOvPd3cZ7MrHtNorW8HRpyrS3d3HmsE0BwOK43muPDBIWUdTbetdX1zo6CqZaoCeDaceXjxceKlnR01Vd7lqO7W663OoqzPTNmBnlc92Wuxwz4OKpR1uhXrxoU0+PWUpavRrVlRgnx6zQY58BhUv0n9Fx33RB1FBE01dmcJN7kTETh4z6j6FdGcBR7aBTRVmir3BO0PjfQT5B7cMJHxhV1+xp6jpta3qLg4v1NLKfqZteg31TTtSo3NN8YyXrTeGvWj8+kRFyDJYbR17F5WQiIqFQiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAuOR8EVpbCdk42jX19Xc2OFntxBnxw61x4iMH4z4edZLStNr6vdQs7ZZlJ8P1b7kY3VNSt9ItJ3ly8RiuP6Jd7ZGNC6C1fq25U8tgsVRUxQzMc+Xd3Y2gOBPlOwM+HNb5gYWxMa7mBhfNbbXbrRSRUFso4qaniaGsjiaGtaPABfbgHK6W2S2Up7L0JRU3Kc8bz5Lh2L1nNW1u1VTaivCbgowhndXN8cc36uo8XDyT5liDV8UkGq7xFKwtc2unyD4vJ/StwHjkLKu37Ss1h1rJd42H2LeB1zHdgkAAe35D6VfbRUZToRqL+V+8jbXaTlRjNdT95XApKs0prhTS+xg/qzNuHcD8Z3d7lnHYpZsw19XaF1HDUMke631T2xVkGeDmE43wPum8/iX27Jbjbqq4VOhdQgvteoGiEccdVUj/ZyNPYezz4XUaz0FeNF6kFjqY3ytmkHsOZreE7ScDHj2Ed61ulTqUowuqL5PD7n8GjAUoVKajcUnyfHufwZr64Xyy2yFk10udJSxTDyHTzNYHcOzJ48F0FgtGzmbUEmo9Mutj7i+N0cj6OdpDmuwTlrTjJxzwoTtu0hXVmz211kLTNPYms6/HEmMsDXu8cEA+bJVJ6P09qG+1U02l6hrbhQsEzI2z9XO8dpj7yO1bNeajKhcRpyp73Jrt9XgbDdX8qNeNN097rXb6jZcVHSQzSzw08bJJyDK9rAHPIGAXEc+HevfgY4qIbLrzfb1pOCbUtJNBcYHvgnE0ZY55aeDsHvBUw58Vm6U41IKceT4mXpzU4KUeTIlftmmkdQ3elvtdaw2tpZmTNmiO4ZC0ggPHJw4dvFd3e7NQ3+1VVmuDHPpqqIxSAHBwe0HsIXY+fCY9SKhTW9iK48+/wASipU1nCXHn3nR6Y0fp/SFF7BsNvjp2Hi52Mvee9zuZXe8cLj0pgYXqEI04qMFhI9xhGC3YrCPnFPTRTy1DYo2SStAkeGgOcByye3GSobaqPZTo24VE1FWWejr55HOlklqWdbvOOSMuORz5DCj+2ifXVyqaDSmkKWoEVZG+Wrmiy0boOA1z+TRxJPHKztW2eso7w+yNdFU1TZhAPY7+sa+QnGGnt4nCwWoal8mqqMaecPm+19hhr7UPQVFGNPOOt9vca+1drKg03pCr1RDPFURMjzTljw5srzwaARz4/IVjq8Xi43+5VF3utS+epqXl73uOfQO4DsC0JtM0nVWfYhSWeFgc+2CnfUBvLn5bv8A7OVc7Kdmgvhfq/U8XU6etrXTO3x/GS0ZIH90Y4nt5Kx1dXF5XhQSxwT7l2t+BZ6mq91WhRSxwT7l2v1Fd1NHV0RjFZSywGaMSxiRhbvsPJwzzBxzVodG+CWXXc0zM7kNFIX+kgD41AdYaim1TqGsvUuWtmfiGMnhHEODGjuAH6VenRu0lLbbNV6orIi19yIip94YPUt5u8xd81Y/S6G/fJQ4qLznuRZadR37xKHFJ5z3IursXU6oop7jpu50FMzelqKSWKNucZc5hAHrK7YDK5zw5LeatNVYOm+TWPM3anUdKamuaefI/PHU2i9U6PqfYuo7LUUTzwDntyx3mcOB9BXTL9Eb9pyz6lt09qvlDFVUs7cOZI0HHiO4jvCxLte2bVGzbVD7a0vkoKnMtHK7m5meLT4tPD1HtXOO2Ow1TZ5fK7eTnSbxl84vv7u86L2O25p7Qz+R3EVCqlnhyku7sfd2cSEIiKOiRgiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIDjwW1OjZaoLdsrtlTG0B9a+aZ7gOJO+5oz6GhYrW3+j1PHNsjsTWOB6tkrHeBEruHxqUuiiEJavUcuag8eaIu6V5yjpFNR5Oaz/tZZAxlRW/7TNE6YrnW69X2GnqWAF0W49zhnlndBUqAA4LOPSP0zNSX6k1PDE4wV0Qglf2NlZyB7st+Qqd9RuKtrQdSkk2u05wv7ipbUXUprOC7bFr7SGpSGWW/0lRIeUe9uvPma7BK92rdJWfWdmks15gL4X+UxzTh8bxyc09hWSNN6S1Hf6eouenGtlmoHAuiim3agdoe1vMjxHatNbIb9eb7o6F+oGyivpJX00xlaWvdungXA9uOas7DUHqEfR1oYyvU11lrZXzvluVoYyvUyp7j0btW26uZUaevVJOyOQPjdIXQyMwcgnGePmV+C001wpKB1+oaWpqqTclDiwOEcwHFzSeXHko7d9sGz+x1klvrL60zxP6uRsUbn7ju0EgYUg0/qWx6oovZ9hucNXCDuuMZ4tPc4HiD5197S3s6E5QotZfNZzy7j7WtG1pTlCk1l81nPI7KWCKeJ0MrGuY9pa5rhkEHmCqyfsE03DqCG/Wa519sdFL1vVU7xgHuaTxaPDirR4cEwBxCvatvSr49JHOORd1KFOtjfWccgxgYMBeS43h3rjfYeTgfSvtlLgfbBwSAM8AvHfZni9ufOqY2l1+1673Se2aYoZ6O2wOMYkgqImyT8BlxcXZaM5wBg8FAW6M2zMPsprbq154dYLiMkjx31d07eE45lUS9Zql7tHXtqzpUrSpJLhndaT8OHE1PkHguSOCqXZRdNqkFc+0a2t8k1I5pdDVyzRGRjh9qd0+UD381bR9K+E4qEt1NPwM7YXny6iqzhKHdJNNHV6ksbdRWaps76yopW1TNx0tO7deB24PiohojYtpfRdaLo181fXMz1c0+MR57WtHAHxVh8Bg5XBGPBWk7elUmqk45a5FxO3pVJqcllrkemppYK2nkpaqJksMrSx8bhkOaRgghQ/aVpS8XzRMmmNJNpKUSFkZY77GwQtOd0YHDkFNs7qjeptoek9IzNpr5d44ZnjeELQXSY78DkPOqXKpODVR4TWM8vaUuFTdNqo8J8M8ip9JdGmVlTHVavukckTCCaWmz5fg557PMr0gp6S20sdPTxRwU9OwMYxoDWsaBwA7gul0ztA0rq98kNhusc80Td98RBa9re/BVb7d67Vt3uVBonTNNVSsqITUVLIGn7IN7DQ48g0YPPhnCsYRttOoOrbx3s9nFt+JZwjQsKLqUFnPZxbJ1ctrezu0VBpqzU9L1jeBEQdKB6WAhdvpzV+ndXRSzaeusdYyAhshYHDdJ5ZBA7isb36yVmnbnLaLhJTuqYMCQQyiRrXEZ3SR2jtHYtJ7AdNz2TRLa6rYWzXWU1IBGCI8YZnz4J/4lbafqdxeXDpygklnPPK/6y3sdRrXVd03FJLnzyiz8ceapDpXWSGs0HBeS1olt1VHh2OJa87pGe7OD6Fd57lUXShmjZssqonOAdJVQbo7yHgr57XU4VNEuVNcN1v1pZXtN82SnOnrds4c9+K9TeH7DGy4OO1cgEnACn+j9hu0LWUTKqitPsWlcARPVu6trh3tHM+pctWOnXepVPRWlNzl3LJ1Hf6laaZT9Ld1FCPa3gr/kuVfrOiHqZ7QX6qoGuI4jqHHHxrqb/wBFfX1riM1qqqK6BoJLYyY3nwAdwPrWdq7Ea9Rg6krZ4XZhvyTyYGltxoFaapxuVl9uUvNrBTC44HgvrutoudjrJLddqGekqYjh8crS1w/6eK+Q47VrFSlOjJwmsNc0zaKdWFaCnTeU+TRyiIvkfUIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiA44ZWmOidraJ9NXaGrJsSQk1VICebDwe0eY4Pp8FmfgOK7zRF3vNj1Xa7lYCfZ0dQxsLB/vC44LD4EEg+dbLspq89F1WlcwWVnDS60+D+JrW1ejw1vSqttJ4aWYt9TXFfDwZ+hY54wur1Bp+2amtk1ou9MJ6ecYc08wewg9hHevtpHzOpYn1DQ2QsBeAcgOxxAPbxX0ce5dX8KsPpLgzlCcFxjLiUna9gN207qaG86f1gaaCJ+8N6EmXd7WHB3XA8uKuhrN1mDjJ5kDGSuvv2orPpmgfc73cIqWBn2zzxce4DmT4BU1qLpNwRTmHTNgM0YOOuqn7mfENGfjWOc7PSk1nGeOOL9hjnUtNNTWcZ6ufsPbtc2K3O8XiTVGk4WTST4dVUhdulzx9s0nhx4ZCmuynSUNgt0lxfpyWx1lWGx1FKanrWEszh4HHBOe9VXR9JrUjJga/T9DNFniI5HNcB24Perq0Nr2xa9tpr7RI5skRDZ4JOD4nHv7x3FWtjOwr3Lq0X9J9TXtRb2crOtXdWk/pPqfvRKkRccCs+ZorLb5qq96U0UKqxVPsaepqWU7pAMuawhxO73HgOKz3ojXms6TVdrLdSXCVtTWQwSxzVD5WPY97WuBa4kcjzV3dJ7hoKlGP/EYvmuWc9JZOqrJkf+I030rVz7t/qd5Q2oo0qVSUYpQ4JtLi+PDvJk2Nsbats/VqVKcW8y4tJvguBS+1Oab65urvsrx/25Wn2x/p3qU6kkk/e7aOd1jsnUFzyd48fsbFFNqXvmau+G636d6lWpP5umjPzgufzGKtOrU9JdfS/lf5kXU4Q9Ha8OtflZ4dGqWU7dtHNMryDWyZG8f6CRfpgF+ZvRp9/jR35ZJ9BIv0zHYpD6PZylp83J/zP3I0bbiKjfQx/h/VnkiIt+NKPFUnta2c1l0rpp9P6QkuNdc3NdJcZKrAgxgbrWZGBgduVdmFwcnkFbXNtC6h6OZb3FvC5huTK92RbMjoG2ST3CRkt0rMGdzOLY2jlG09veT2lSzUlsr7tZqqitVxNvq5oyxlSGbxZnu/Wu3yB2oCCO9Vp29OlS9DBYR6p0IUqXo4rgUppfo5UVDcG3DVF1FybG/fEDGbrHnOfLJ4kZ5jtV0RsZFGI42hrWgBoAwAO5ezn4pgKlva0rVNUo4yeaFtStlimsHBI4FZd6WOs4qq4W/RdJI1/sT+FVWD7V7hhjfUSfSFd21PaNb9m2m5LtVtMtRNmKkhH+8lxwyewDmVlrZNYqzattUZXX15qGNkdcK1zhwcAeDfMTgeZR7t5qruPR7P2jzVrNKXdHPX4+4k7YXS1buptBeLFKim1/mljq8PfgtLYLsJoqahp9ZaxomzVUwElJSyDLYm82vcDzceYB5edaEZG2NobG1rWgYACRRsia2JjQ1rRgAdgXpuNxorRRS3C4VLIKaBpfJI84a1o5krb9G0az2es1QopJJfSl1t9bb/AO4NQ1jWLvaC8dxXbbb4LqS6kl/3J9PHsTCq6DpFbOai5C3trKljHP3G1D4CIic888wPHCsuCpiqomTwSNfE9oc1zTkEHlgq6sNWsdU3lZ1Yz3eDw08Fpeadd6fu/Kqcob3LKayQzabsqsO0i0yU1ZA2GvjaTTVbW+VG7sHiO8LE2qNNXTSN8qtP3iDq6mlfuu7nDscO8EcV+iZ86oDpW6IgrdP0+taaPFTbntgnIHtonnAz5nEeglR/0ibLUb2zlqdvHFSCy8fzJc8965kg9He1NawvI6ZcSzSqPCz/ACyfLHc+WPWZXREXPJ0MEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAFPthFBBcNq1hgqG7zWTPlA/vMY5zfjAUAHMqUbMdQs0truy3yR27FDUtbKcZwx3ku+IlZbQ6lOjqVCpV+qpxb8MoxOu0519MuKdL6zhJLxwb/AGjyQvgvtdVWu0VddRUElbPBC6SOnjOHSOA4NC+uCVk8TJo3hzXgOaQcgg8l7POuvvrw+izkGUXxXIxPrLV1/wBYXeWvv1Q8va4tZBxDIAD7Vrezx7V1FDTMrKyGlkqoqZsrwwzS53GA9pxxwtLbT9h1Bq2WW92B0dFdHDL2EYjnPefuXePb2rP1/wBEaq0xO6C82Sqg3f8AeCMujPiHjgtCv7C5oVXOonJPr7fgaTe2VxQqOVRby7e34EptuyO3V+4frmacAceQl8r1Owrk2Z7IqDQ1e69UWpaiudNCY3MYA2F7TggkAnOOzzlZap6SqrJhT0dNLPM44DImFzifMOK0bsK2bXzTYl1FqF81PLUR9XBSGQ+Sw8SXt5Z4DA7Ff6Q4VKy3aXFdeXwL3S3CdVbtLl15fAuXOeS4xjsworfdpmi9OV31Out/gjqB7ZjcvLPwt3l5uaktPURVcEdTA/ejmYHsOCMtIyDgrbYVYTbjFptc+42eNSM21FptFSdJ/wBwlN8IRfNcs5aR91Vk+EqX6Vq0d0nQDoSn48q+I/E5Zx0j7q7J8IU30jVzp0i/ayj/AKPeTlsT9m6vjP3FK7U/fM1f8OV3071KtSfzdNGfnBc/mMUV2o++Zq74brvp3qU6k/m66M/OC5fMYrun+8uvwv8AMis/3dr4r8rPX0ac/X30bj+uyfQSL9MvBfmb0as/X30bjn7Nk+gkX6N6k1Fa9KWeovl5qBFT07cuPaT2ADtJPIKQNgq0LfS6tWq0oqTbb5JJLLNI21pzrajTp01mTikkubeWdhVVVPRQOqamZkUUbS573uDWtA5kk8lWOpOkRoaxyup7dJNdpmcD7Hbhmc8Rvnh6lRG0XatqHaDWPEsz6W1sfmCjY7gB2OeR7Z3xDsUWtNmu17qm0Nmt1RWTO+0ijLseJ7APErT9d6VLmtcO00KnnjhSay34L/vgbFpHR5QpUflGrzx1uKeEvF/DzLtrelTVGT+AaTb1f/rVPlf8owvT++pu/wB6VL/mnfsqHUewTahWxda2xxQ9m7PUsY71DK+j97xtR/s6h/zY/UsH+1+kCv8ATjGok/8AIl+hlf2ZsXT+g5Q4f538SfWjbZdNfWTUdC60MtzqS1yVLJoqhxcHDgMcBjvyo1s/6Rd8swht+ro33KjADRUtAE7B/e7H/EfOvp0tss1loezaouWoqSnigltEkTDFP1hLufIeAUJ0Fsg1Zrl0c0FM6htxI3quoaWgjh7RvN3A8+Xir241Ha6M7N0975RJS3otYylLhvLCWMdZaULHZuULpTcfQJxw084eOO68t57vYaz01qyw6tohX2C4xVcPJ26fKYe5zTxB867gZ8FCtn2yrTuzyEutrHzVkrQ2aqkPlP8AADk0eA+NTbBHIKdNMneVLWEr+KjVx9JReVkiW/jawuJRs5OVPqbWH/3/ALgrzbvpqn1Js2usckYdPRxeyYHY4te3j8Yyqs6H9LA52oq10Y65nURh2OIad4kesBXXtRro6DZ/fKmbkKN7ABzLnDdAHeclQbo27ObzojT1Xcr1iOovJilFPjyoWNDsb3id7l2LUNQ051trLa6pwzuwbk+xYaWe9t8DbtP1JUdlLq1qTxvTjurteU5Y7klx8S5M55qg+lJfKynobTp+F5bT1jpJ5sH225gNae8Zdn0BX5hUx0j9FXPUFnor7aaaWpltReJIYxlxifjJA5kgtHAdmVf7d0bqtoFxG0TcsLgubWVn2ZMVshUt6WsUZXLSjnr5Zw8e3BmPBxxWoOjRfq25aQqrZWymRtrqeqgJ+1jc0ODc+B3vMMLMcNNVVU7KSmglknkcGMjYwue53cAOOVrzYloWp0No9sFxBbX18hqqhmciNxAAb6GgZ8cqGuiizvXrDrQTVOKal2ceS8ckn9I1zbLS1Sm1vuS3V18Ob8McPWWGRlQ7a3bZrrs5v1FTUr6iZ9HJ1UTGFznPAy3AHEnKmPcjsHgePgujLq2jdUJ0JcFNNeawQla13a14V484tPyeTCEOxLapURtmj0XXFrhkbzmNPqLsrz+sZtZ+8qs/xI/2luwNaOwLnA7lGS6JtLfOrP2fAkv519V6qUPb8TCX1jNrP3lVn+JH+0n1jNrP3lVn+JH+0t2brO5N1ncq/NLpX3s/NfAr87Gq/dQ8n8TCf1jNrP3lVn+JH+0n1jNrP3lVn+JH+0t2brO4LjdZ3J80ulfez818Cnzr6r91DyfxMFV+x3abbIDU1ejLgIxzLGiQ+ppJUWrbfX26Y01xopqaYc45Yyxw9BX6Pbjc+1C6i/aT09qakfRXu001XE8EESRjI8QeYPiFYXvRJQlFuzrtPskk15rGC+sulq4jJK8oJrti2n5POfYfnhjhhCro207A6jRDZNR6YbLU2fOZYj5T6bPb/eb49naqX8VEOr6Pd6Jcu1vI4kvJrtT60S9o+s2muWyurSWYvzT7GupnKIixJlgiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiLgU5mmtgG3SilooNF6urBBUQgRUdVK7DZG8gxxPJw5AnmtFtc14y05BX5sgkHIOCFdeyzpIXfSkUVk1UyW425mGMnBzNC3u4+3A8eKmnY3pEhRpxsNVfBcIz7ux/EhfbLo7nUqSv8ASVz4yh39bj8PI16CMLwfGx4w9gI8RlR3Se0LSGtKYVGn7zBUHGXRl27I3ztPEKR7wdyIUzULmjd01UoyUovk08ohqvbVbabpV4uMlzTWH7T0xUlJE4ujpomu7wwAqObTNUSaQ0bcLzTj7O1gih8JHndB9Gc+hSkjIx2rq9Qabs2qLe61XyjbU0pe2TccSPKHI5Bz2r3WhJ0pKnwk1w8S1qQcqco0+DxwMd6f1Ayz3sX64W9l0nYXStbUvODKTkPd90QeOCtQ7J7nq296cdedWFzZK2dz6eLqhH1cPIYHPB5jPFfRatlGgbJVCsodN03WtILXSl0u6RyIDiQD4qXBoAAHLuCxWmabWs25VJ5XYvezG6fYVbV5qS9S97Kh6TnuDg/L4vkcs5aTOdV2X4RpvpGrR3Sb9wcB/wDfxfNcs46U91dk+EaX6VqgrpD+1lLwh7zoXYn7OVfGfuRSm1H3zNXfDdd9O9SnUf8AN20b+cFz+jYottS987V3w3XfTvUp1F/N10b+cNy+YxXdP95dfhf5kep/u7XxX5Wevo0nG3fRx7qyT6CRaS6R2tp7rqRuk6WZwo7Y1r5mg8HzOGRn8EEetZq6N79zbnpB3dVy/QSKf3mtkuN5r6+YnrKmqlld53PJ/SsXq+r1bHZ35FSePS1Hl9ySyvW2i8sdNp3eu/Kaiz6KCx4tvj6kmdzs/wBDXDX+oI7NQHq4mjrKmY8oos4J8SeQHetf6R0XY9F2qO12WkZExoG+/GXyO7XOd2kqB9G/TMVr0R9WzuunusrpC7uYwlrR6w4+lW1k8MKTejnZa30rTqd9UjmtUW9nsT5JeriyPNt9oa2o307SEsUqbxjta5t9vHghheqWeCnY6SeVjGtGSXHGB51ANqm1q27O6MU8TG1d3qGkwU+9gNHLfeexvxlZh1TrzVWsat1VfbxLK0nLYWOLImeDWDh28zx8VdbU9IWn7NzdvGPpKvXFPCXi+rwxkt9n9i73XIKu3uUnybWW/Bfrk1leNpWzSGJ9HdtTWuSOQFr4+sEgcO4hufjX2WbXmhrmI4bRqG2yOdwZGyZod5g04PxLEOBngOK5b5J3gS1w7RzCj6HS/eqr6SdrHd7s58/+DdZdGdq6e5G4lnwWPL/k/QBsjJBljgfMVz4LG+g9sWrNE1MTG1klwtocOspJ3l2G/wBxx4tPxeC1dpPVNp1lZoL5Z5hJDOOR4OY4c2uHYQVKOy22lhtTFxo5jUXFxfPxT60R/tBstebPSTq/Sg+Ul7mupnYXC20V0ibTXCliqIg9rwyRoc3eactOD2gjK+oANGGjgF5AcUIW3qMU3JLizWm20k3wR65JGRML3uDWjiS44AXgyaGdgcyVjmntDgQs5dJrUt3bfaLTUNS+KiFOKp7GOI61xc4De7wN3h51T1Ffr7bc/U29V9KXcCIah7M+oqLNd6UKGi6lOwlQc1Hg3lLj14WOXrJB0jYCvqthG8jWUXLiljPDvef0NwC22Cjq33FtJRQzyYD5txrXOxyBdzXYskZIwPYQWnkRyWC6y73e4P62vulXUu5Zmnc8+slay2BzTzbL7U+eV8jg+ZgLjkholcAPMArzZDbuhtJfTs6Nv6NJOWcrjxS4pJcXktdpdkKuhWcLqrW322o4xy4N88vsLGBBQ4HEp4r4bzNNT2qqmpx9lZDI5nDPlBpI4edSNVmqUHN9RpMI78lFdZ8Vx1lpW1VHsS5aht9PMOcctSxrh5wTkLzp9X6YrG71Nf7fKP7lQw/pWGKueprqmaqrnufUSvc+RzuZcTx+NescB5JI8ygmp0x1o1XFWy3c8OLzjq6iXYdF9KVNN3Dz1/R4e83kNQ2M8rrSf47f1rn90Fk/tWk/xm/rWDN533bvWud539I71p881T/1V/u/4HzXQ/8AYf8At/5N5C/2U8rpSf4zf1p9XrN/alL/AIzf1rBu87+kd6033Hhvu9ar880//UX+7/gfNdD/ANh/7f8Ak33DUwTjMMzJPFpBXtAAHisOaN1Xe9L3+ir7XcJmYmY2SIyHckYTgtc3kRgnzLcMR3o2uPaMqRdjtr6W1tGpONNwlBpNZyuPLD4dho+0+zNTZurCDnvxknh4w+GM5WX2nrraOnuFJNRVUbZIpmGN7XDIcCMEFYL2p6Tj0Tr266epyTBBKHw57I3tD2j0B2PQt9kHsKxR0k+O127fi6f6Fq17pXtaU9Mp3DX01JJPuaeV7Dauim6qw1Spbp/QlBtrvTWH7WVgiIufDoIIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiA9lNV1NHMyppJ5IZWEOa+Nxa4EciCFO7Ft12nWSSEN1RU1MEcjXPiqA2TfaDxaXOBcAeXAqALjj3LIWeqXuny3rWrKPg2jHXulWWoLF1SjLxSZvzZ/tCsu0GyR3a0zt3w0CeEu8uJ/a1w+Q9qlYHavzt07qi/6TuLLpp65z0VQz7aN3Bw7nDk4eBVw2Ppaatooty+WOjuLgAA9jzCT4kcR6sKcNC6ULGtRjT1NOFRc5JZT7+HFZ8CEdc6ML6hWlU0xqdN8k3iS7uPB+OTWPD1pkYzjgqp2Nba6jarXXKjmsLLeKCOOQObUGTe3iRj2oxyVrN4jJUkabqVtq1tG7tJb0JZw8Ncnh8GRxqWm3Ok3ErW7juzjjKynzWeayVD0nPcHB+XxfI5Zv0p7qrH8I0v0rVo/pPEjQdNjtuEXyOWcdKe6myfCVL9Kxc/9Iv2so/6PeTDsR9nKvjP3FKbUffM1d8N13071KdRfzddG/nDcvmMUW2on/vM1d8N13071KdR/zddH/nDcvmMV3T/eXX4X+ZHqf7u18V+VnzdHL38NI/lc3+nlU6qv4xN+G75SoN0c/fw0h+VTf6eVTmq/jE34bvlK1DaT+G2/45+6BsGjf3+t+CHvmbC2Ge9ZYT3RSfSuU0rqqKhpJquZwayFjpHEnAAAySoXsM96yxfipPpXLvtcsfJo68sjaXOdQTgADJJ3HLpHR6jo6BQqR5qlF+UUQPqkFU1itB8nUkvOTMY6rv8AVaq1JcL/AFjyX1k7ntbn2rBwY0eAaAFMtimzW27QLtVvu9Tu0dvDHPgjdh0rnE4BPMN4HJCrZvip5sSvtXY9otsEEjuqrXmmnYMkOY4HHDvBAK5h2fr0LraCnU1KPpIynxT623zfbhvkT/rVGtb6NUp2EtyUIcMdSS5L1LCZqix6I0rpyAwWaxUdMHAbxZEMux2kniT4ldXrTZVpHWVHNFW2yCGqezdjq4WBssZ7CCOYz2HgpkCDyTj3Lq6rpNjXt3bTox3OWMLHq7DnSnqF3RrfKI1JKfPOXn/kwfqOwVulr5WWC4DE1HKYy7GA8c2uHgRg+lXh0V6w9Vfbe+pJw+KaOIu5ZDg5wHj5IJ8Aoh0j44m7Rt6MAF9FEXYGOOXevgux6MEhGs7nH91Q5/5wuednLeOibb/JKLzFSlFeDTx5E1a5Xlq2ybuan1nGMn4prJqFERdMkEkV1ns50vrqOFuoKEyvp89XKx5Y9ueYyOY8Cqc2q7CLDprTFTqTTMtU19EA+WKV4cx0ecOI4ZBGcrRZB4cVC9r4xs11AO+jeMLT9qNnNLv7K4ua9GLqbsnvY45SeHnuNk2f1zULK6o0KVWW5vRW7nhhvjwMY8MYWw9hUUcWy6ybjcb8Tnu8XFxJKx6tjbD/AHrrF+IPylRD0PRX7XrfgfvRJnSY/wD6yl+NflZO8ccr11MbZIJGO5OaR6wvbkL47rOKa3VVQc4jhe4458GkromtJRpycuSRCNNOU0lzMJXdnV3euY0cGVUrR5g8q3NjGxW36utg1NqfrTRvkLaaCN+4JWjgXOI44znAGOSp2rnFTVTVLc4mkdJx58XErYmxNoGzGxDA/i+fjK5p6O9Gs9b1ys7qClCCclF8m8pLPgTxtxqd1pWk0/k8t2Umk2uaWM8OzkehmwXZa3idMNce908h/wD0ufrDbLPvWj/xpP2lYPiviulzpLPb57jXTNhgp43SSPdya0DJKnyez2i0oOc7amkv8q+BDcdb1WclGNxNt/5pfEq/WGzXYtouyT3m66egYyJpLGdc/fkf2NaN7iSsv1s8FXXT1NNRx0sMkhMcDCS1jSeDRnjwUq2nbQqzaHqGSuL5GW+DMdHTuPBrPuyPuncz6B2LrNE6PumuL/T2O2RHyyHTzY8mKMHynH9A7Suc9q9Rt9odTjYaNbxjFPdW6knJ8s8EuBOGztlX0SwlearWlKTWXvSbUVzxxfPt7+BIdjez+t1lqulqZqR/1KoJBPUSvYdx5aciMHkSTjPhlbCa0MYGDkAup0vpm16RslPY7PAI4KduP7z3drnHtJPEldv25U9bGbL09mLD0LeakuMn345LuRDu1Gvz2gvPTYxCPCK7u197Oc9yxL0kHtftcvG64HdZTtODyPUtWx9RXyg05Zay9XGVsdPSROke4nsA5ecrAGq9QVGqdR3G/wBWcyV07pPwW58kegAD0LVeli/pQsKVkn9KUt7wSTXtbNz6KLCrPUKt7j6EY7vrbT9iXH1HVoiKAifAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiA0H0P/AOXNRfk1P8561IOZWW+h/wDy5qL8mp/nPWpBzK6d6OPs9R8ZfmZzH0i/aGt/p/Kinek/7hKb4Qi+a5Zz0n7qrJ8I0v0rVozpP+4Sm+EIvmuWctJ+6qyfCVL9K1RV0i/ayj4Q95vGxP2cq+M/cUrtR98zV3w3XfTvUo1Fx6Ouj/DUNy+jjUX2o++Zq74brvp3qU6h/m6aQ/OK5/MjV3T/AHl1+F/mRWX7u18V+WR8/Rz9/DSH5VN/p5VOar+MTfhu+UqDdHP38NIflU3+nlU5qv4xN+G75StQ2l/htv8Ajn7oGwaN/f634Ie+ZsLYX71ti/FSfSOU6ljZLG6NwyHDBHeFBdhfvW2L8VJ9I5T08l0zs6k9GtU/u4flRAWtPGp3DX+OX5mZq1n0bL4y7T1ekKmmmop3l7IJnlr4snJbk8C0dnb8q7zZPsEuum7/AE+p9U1MHW0eXU9NCS7DyCN5zuXAHgBnmra1PrTTejqT2XqG6RUrT7VpOXu/BaOJVP6g6UdKwywab09JMeLWz1Mm43wO4Mk+Y4Wgaho+xuzeofL7meKie8o5bSfPKSWVx7eBuVlqm1Ou2fyOhHeg1hyxjK5Y3nw5dnEv0YAxwC+W5XOgtNHLXXGqip4IWl8kkjt1rQO0lZRrukNtNrciO4UdI08xDSj5XFxUMvur9S6nkMl9vlVVgnIZJIdwHwaPJHoC+WodL+mUoNWdKU5dWcJe9v2H0sujS/qzTuqkYx68Zb9yXtOy2latbrXWVwvsJPsZzhFTA8+rYMA+GTl2PFTfow8dbV5/9gfntVQAjkrf6MXu3r/yA/Paou2QvKmo7V0bqt9ac234tNkgbS2tOy2cq29L6sYpLwWEalRcBcrrA52OP+ihu14Z2bX/API3qZf9FDtrozs2v/5E9YrXf4Zcfgl7mX+lf36j+OPvRi3PDK2PsO96+xfiD84rHDeXqWxthvvX2L8R+kqBuh/+LVvwfqiYOkz+GUvxr8rJ7zC+W4Uja2inpHOIE8boyRzw4EfpX055JjGV0XOCqRcZcmQjGTi00Y9uOwraLR3h9rpbG+phLyIqlr29WW5OCT2cOa1DoPTs2lNI2uw1Eoklo4GskcORd249KkGBnmuSM8lqOzuxdhs3cVbm1cm59r4JZzhYRsmt7U3uu0KdC4SSjx4J8XjGXls4cQ0FxOAsv7fNqJ1HXv0hZZ3C3UcmKp4PCeVp4AY5taR6T5lPNvm1Iabt7tJ2Sci510Z62VjsGmiPb+E7iB3c+5Zh48+3mo/6Ttst1PRbGXF/Xa/L8Tc9gdl95rVbtcF9RP8AN8PPsPooaGsuNZDbqCnfPU1DxHExoyXuJwAtf7J9nFJs+0/HFLFG+6VQD62ZvHJ7Gg/ctz8pUM2A7KjZqVmtL/SAV9Q3NJG/iYYyPbY7HO+Iecq7+XYsr0a7GfsuitUvI/2s19FP+VPr8X7jHbd7U/tCq9PtZf2cX9Jr+Zr9F7/UeS8XENaXHgAvJcEBwweRUuEcGQ+kTtdOqrm7SViqP+y6GQid7TwnlHD1N7PH0Kk/ELTnSJ2K0ktFPrzTFL1dTCDJXwMGGys7ZAOxw7e8LMfFcubd2+o0NXnLUHne4xa5bvVjsx1nT2wlxp1bR4R09Y3eEk+e9157c9XccoiLSjdQiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiA0H0P/AOXNRfk1P8561IOZWW+h/wDy5qL8mp/nPWpBzK6d6OPs9R8ZfmZzH0i/aGt4R/Kinek/7hKb4Qi+a5Zz0j7qrJ8I030rVozpP+4Sm+EIvmuWc9Ie6ux/CFN9I1RV0i/ayj4Q95vGxX2bq+M/cUptR983V3w5XfTvUp1D/N00h+cVz+ZGottR983V3w5XfTvUo1B/N10h+cNy+jjV3T/eXX4X+ZHqX7u18V+Vno6OXv4aR/K5v9NKpzVfxib8N3ylQXo5e/hpH8rm/wBPKp1Vfxib8N3ylahtL/Drf8c/dAz+j/3+t+CHvmbB2G4OyyxA/wBFJ9K5dvtC1fTaH0tWX+cbz427sMf9JIeDW+vifAFdRsMI+tbYs/0Un0jlXXSmu8oZY7GzIY90tS/Due7gAEenKna51eWibHwvIfWVKCXi0kvIh6302OrbTTtZ/VdSTfgm2/PGCkNQahu+qbrNd73WPnqZSeLjwYOxrR2Ady92nNI6k1dVGksVomqi0gOe0YY38Jx4BfPYbNUagvlDY6U4lrZ2RB2M7oJ4n0DJW2dK6XtekrLT2a1UzY4oGAE4G893aSe0kqGdjdka22lxUuryo1TT+k+bbfHCfvJQ2o2lp7L0KdtawTnJcFyUUuvh7EZ1tPRm1tVmN1zr6ChYT5QDnSPA9AwfWppZei7p6mlZPeb7WVu64ExRsbExw7jzd6iFeHLwTJxzU0WPRvs9ZYfod9rrk2/Zy9hFt1txrd1lel3U/wDCkvbz9pkzb1pbT+kb5bLbp+2spIjSlzw0HyjnGS48XHxK7Howgfu3r8/2efnherpM1T5td0tMcBsFCzHnc52fkC9nRiz+7e4Z/qB+e1RVb06VHb9U6EVGKnhJLCWF1JEh1Z1auxjnVk3Jwy23lvMu01E4tHE8Fw14PJwdjtVNdJbU10sunrfbLVXS0xuE7hK6Jxa5zGtzu5HEAkjKhfRy1peW6rfpmtr6ippK2Bz42SyFwikYM5GeWRkEDwUs3e21raa7DQ5we9LC3s8E3xSwRzb7J3Fzo8tXjJYWXu44tJ4byacPYojtYbvbOL+MZ/gUnyKXHsUS2ru3NnN+dnGKKT5Fset/w2vn/BL3MwWmf32j+KPvRike19S2NsO96+xfiD8pWOR7X1LY2w73r7F+IPylQN0P/wAXr/g/VExdJn8MpfjX5WT5ERdHEIHiobtN2g0ez7Tslyk3ZauXLKSDODJIeXoHMnuUhvt6t+n7TUXi6VLYaamYZHvceQHZ5zywsabQ9dXDX+opbzW/Y4I8xUsOciOLPAec8yVoO3u10Nm7Hcov+3nwiuxdbf6d5uGx+zU9du9+qv7KH1u99SXj19iOiudzrbzcai63GYzVNXI6WV57XHu8FaOwrZWdWXNupr3TB1oon5jjeOFRIOXna08T3nh3qH7ONA3HaDqBlspQWUkGJKyfsjjzyHe48gPT2LZVmtFBYrbT2m10zIKamYGRsaMAAKMOjrZKet3T1jUVmmnlZ/mlnn3pe033bbaSGk2/7MsnibWHj+WP6N9XYvUfYGiMBrQABwC6TWGsLPoiyy3u9SuZDGQ0NYN573nk1o7Su9I9KovpTTSix2WnBPVvqXvd3ZDOHylTbtNqk9E0mte0VmUFwzyy2kiKdA0+Oq6lStKjxGT4454Sy/cSfQm3TTWubz9QoaSpoaqRpdA2fdIlAGSAQeBAGcKzOxYy2Mxvk2nWFkTt1wmcc+AjcSPSAVswHgB4LBdH20N5tHpsri9w5Rk1lLGVhPl3ZMxtpotrod/GjaZ3ZRTw3nDy1+h66qCOqp5KeVocyRhY4HkQRghfn3ruxHTWsbxYsANpKt7GAdjCctHqIX6FcMLCO2852rakI/rQ+Y1YDpaoQdjQrNfSUseprj7jaOiavNX9ejn6Lhn1ppL3sg6IigMnwIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgNB9D/wDlzUX5NT/OetSDmVlvof8A8uai/Jqf5z1qQcyunejj7PUfGX5mcx9Iv2hreEfyop3pP+4Sm+EIvmuWc9Ie6ux/CFN9I1aM6T/uEpvhCL5rlnPSPursfwjTfSNUVdIv2so/6PebxsV9m6vjP3FKbUffN1d8OV3071KdRfzdNH/nDcvmRqLbUffN1d8OV3071KdRfzdNH/nDcfmMV1T/AHt1+F/mR6n+7tfFflZ83Ry9/DSP5XN/p5VOqr+MTfhu+UqDdHL38NI/lc3+mlU5qv4xN+G75StR2l/h1v8Ajn7oGf0f+/1vwQ98zYOwwf8AdZYvxUn0jlUvSlikZqWzyucSySmkDR3EOGflCtvYZ71th/FSfSuUT6TWmX3HS9LqGnj3n2qf7IQ3J6p/AnPYAcKY9obOpe7DQhSWXGnCXqSTfsIs0W6habWylU5OpNettpe3BSmx+SKPadYHzkbvskgEnGCWOAW0WkEDHbyWBKStnoKyGuo5Syop5GzROH2rmnIPrC2Bs42qaf1zaoHGsip7mxgbUUsjw14dyJb3tPMYWudEWtWtCjV02rJRm5byzwzwSa8VgzvSVpdxVq0r+nFuCjuvHVxym+555k8458AuJHhrS4r1yVMELDJLKxrQMkl2AB4qkNsm26gpKKbS+kKxlTV1AdHUVUZy2BvJzQe1548uAUsa3r9loVrK6uZpYXBdbfYkRzpOkXWsXEbe3jnPN9SXa2VDtZ1JHqjaBdLlTvLqeOQU0JzkbkY3cjwLt4+lTDoxZ/dtcMf1A/PaqfHNWl0dbxS2rXr4ap7WCvpnQsc5wA3gQ7HHvxhc07Man8r2rpX1w8b82/POETttBY/JtnalpQWVCCS8Fj4ZJj0qjmKwMJ/3kxHqCgnR899Gg/Ez/MK7rpL6lorrqWgtFFK2Q2yF5mcxwc0OkI8nh2gNyfwgvi6NtC6r2hmrBGKOjkkdnt3sN/Stk1KrG/6QIOi8pTguHclny4mEsYStNi5elWMwl/8As3jzyjVvgohtdIbs31Bntonru79qaw6Yp21d+utPRRPOGuleBvHwVW7XtreiazQtxtNlvVPX1dfH7HYyF2S0E8XHuACmnaTVbK10+4pVqsYy3JcG1ninjhz4kVaJp91c3lGdKnJx3o8Unjg1njy4GY+7itjbDeOy6x4PKFwPn3isc8C09i0XsO2qaQs+i4tP3+7Q2+ooZJABO7Aka9xcC0+GcEeCgnor1C2sNYqO4moKUGk20lnKeMvuJe6Q7Kvd6ZBUIOTjJNpLLxhrOF3svgcOJXi54Y0uc4AN4knsXT2DV2nNUMkdp68U1b1Rw/qnglvnUQ28arqNL6EqGUMnV1dxe2ljcObQ7O8R3HdBwe9dDX2rW9nYVNRUlKEYt5TTTx1Z73wIVtNOrXV5Cy3XGcmlxWMZ7im9uu1B+rru7TtnnzaKCTynsJ/hEo5k97Wnl6+5VvZLLctR3alsdppjNVVTwyNvZ4k9wA4k9wXxAEBXjsQu+zPRdA69X3UVELzVtxuuad6mj+4BxzPMkeA7FzDaOe22uuvqVVQg3lttJKK5JZfq9pP1yo7KaQqVjSc5JYSSbbk/5njzfki5dnWhbdoHT0VppmtdO4CSqmA4yyEcT5uwDuUs4qC/Xv2YffZS+p36lx9e7ZeeB1ZS+p36l0jZ6todjQhbUK9OMIrCSkuCXrIMudN1e7qyr1qM3KTy3uy+BOhnOPjVK9KOFrtJWyY+2bXho9Mbv1KUT7e9mNON46jEn4uF7/kConbLtXj2iV1PRWmOSG10TnOaZODpnnhvlvYAOA7eJWp7d7U6S9FrW9OtGc5rCjFpvOVxeM8uZsWyGz+pLVqNedKUIweW5JpcuXHtPl2EQvl2pWdzWFwi657jjkOrcM/H8a2GeYHesj9Hqpjg2m0cb+c8EsbPPu73yNK1vyXx6IoxWhyaeczee7gj7dJMm9XimuUF72cu5FYQ23e+vqP8pHzGrd59qVg/bYHDapqPfOT7L/8Ay1eOlr+GUfx/ozJdE38Urfg/VEJREXPx0AEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQGg+h/8Ay5qL8mp/nPWpB3dyyx0QZ426hv8AA54D30sLmt7SA52flC1OO9dOdG7T2epeMvzM5k6RU1tDW8I/lRTnSf8AcHS4/tGL5rlnPSOf3WWMf/0aX6Rq13tV0TTa60tJbKm4OoeoeKlkwZvhrmg+2bzIwTyWf7Jo/Quk7xT3y/7Q7fcW0M4mho7a1zpJXNOWbxOA0ZAz2eK1LbDY/WNZ2jpXdlS3qf0cvKwsPjnrMjoG2GjaFoVW21CuoTe80nzeVwx2mWtqPvm6v+HK76d6lWoyP3uujvzhuPzGKeax2EaT1nqm4assW16zW+mvFZLWz0l0ieyemMjt5zRu5D8OJxy4dp5ru7nsw2ZXnZ7a9l9v2kOpKqy1c1c27VVJilqpZAA9u6DvMAAbg/LlXkNjtZ9Jcv0L4ppcVx4p8PUi2nt9s4qdr/8AKjzXXy+i1x7OJT/Ry9/DSOf61N/ppVOaofwiY/33fKVINlmyXQuy7VVDrnUm0yhvdTbnPdSUVnjcQXlpZvPe/GBhx4YHHtUjZs00pfLn1lj2nWkUlRI4iKqjeypjBPBu5yceIGcjK1rXdgteuNOoxpUG2pSbWVlJqOG+PcZzSukTZmnqNVSu48Yxxx4PDk2l5ov/AGGe9ZYc/wBFJ9K9TK526ku1BPba2FstPUsdHIxwyHNIwQuq0RpmLR+lrfp2KpdUNo493rS3dLySSTjs4k8F3xHElTZpFpKhpdC1uFxjCMZLnxSSaIy1G5jWv6txRfCU5ST5cG20zGu03ZXedn9zkcIJKm0yOLqerDchrfuJMe1cPUefgoQx8kbhLHI5jhxDmnBHpW+6qmpq2J9NVwMmjeMOY9oII8QVALvsF2bXeR05shpZD/VpnRDPfug4UQa/0TVZ3ErjSKiSbzuvKx4NdRJmjdI9OFFUdTg21w3lh58U8cfAyY+6XWaPq5rnVuZjBa6d5B9GV7LLYrvqGvjtlkoJaqokIAZG3OB3k8gPErTtD0bdndM4uq2V1Xxy0SVJaB4eTjPpU/sGktO6XgNNYbRTUbDz6tgyfOeZVhp3RNqdzVUtTrJQ60m2/BZ4LxLu96R7ChTasKTcurKSXxZRd+6Pxs2zaWppmezNQwltVM5mcFo9vGwdoAz5yPMFRIc9jt5rnNc08COBB/Qv0CwCMcwq91ZsP0JqupdXTUD6KqfxdLSu6vePeW+1J8cLP7VdF8btU6ujYg4RSafDOOTz29phtnekCVq5w1TMlJ5TXHGerHZ2Y5GPy5zsulJc7mSTknzrS/Rr0bV2ax1mpLjSmGS6Oa2nD24d1Lftu8BxJ9AB7VJLDsC2d2SZlSba+umZgh1XKZBkdu77XPoViRRRxMEUbAGgYAAwAF9diOjqvol4tS1Gac4p7sVl4b4NtteR89q9t6OrWvyGyi1FtbzfDguOEs9pmPpPGu/djbxN1nsX2F9hz7Tf3zv48cbmfQqcJAPEjh4rd190zYNT04pb9aqatiYctEzA7dPeD2LrLbs00Jan9ZQaYoI3/ddSCfjVvtH0Y3mtarUvYV0oTeeKba4JY7PA++h7f22k6dC0lRblFY4NYfHn3GJd4AZ4YQOZ2H41vE6dsThumz0eO7qW/qXT1+zDQNxd1lZpa3yO48epAPHzLDVehq6iv7K5Tfemv1MnT6UKDf8AaW7S7mn+iM79HEXJ20RvsHe9j+xZPZZHLc+1z/x4VsdI2wXC96JjqbdTSTut1S2okjY3ecWbrmkgDjw3s+bKsDT+ktO6WifHYLPT0LZSDJ1TAC8+J7V27mNeC1wBaewqQ9H2MlZbP1NFuam855y1yWeWM9SwaVqe1Cu9ahq1CnhRxhPm8duO3J+f2QOGB3JvAdvxrbldsz0Fcp3VVdpS3SzO5vMIyV6G7JdnDTkaOtmfxAUby6HL9Se7cRx4M3mPSfaYW9QlnxRisOYeR+NMs+7HrW4Rs90S1rGDTFuwwYaPY7eA9S9rdD6QaN0acoAPxDf1L6Loau+u5Xk/ieH0oW/Vbv8A3L4GGN5naVzlvYQtxu0Doxzg86at+8OREDf1L4JdkmziVxe7SFtLnHJPUgLxPobvucLiL8U0eo9J9o/rUJL1p/Azl0fbdV120yiqqZhdFQRyTTu7GtLC0eklw+Na67crq7Jpmw6chdBYrVTUUbjlzYWBu9512qlfY3Zt7Lad8knPek25Nrll4XD1IjvafXf/ACC9VzGG7FJRSfPhxy/Wzg8QcrCW2852rajI/rQ+Y1bjuldBbbfUV9TI2OGnidI9zjgAAZJX586tvT9R6lud8kJPsypkkbnnuk+SPVhaf0tXNONlQt8/Scm/Ulj9Td+ia3nK+r3GPoqKXrbT/Q6pERQKT0EREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQE02P63Og9dUN3meW0spNPVYGT1b8ZPoIafQt3QzR1ETJo3BzJGhzXDkQeRX5uYWy+jltCZq/R8VorJg642ZrYJATxdH9o71DHoUzdFWvKnOek1nwfGPj1r9SGelPQt+MNWorivoz8Op/p5FtlocHNI5jCyXtp0G7RmqH1NHGRbbmXTwcyGPz5bPQTkeBWtcjkFEtpui6fXWlam0PIbVNHXUsmOLJW8vQeR8Cp9s6/oKmXyfM5r2q0ZaxYSjFfTjxj+q9ZjZF51FNUUdRLSVcL4Z4HujkjeMFjgcEHzFeC2VEASUoy3ZBW30etCy3u/O1VXRfwK2HdhDhwknI/8AyOPnIVYWa0V2oLtSWW2xh9VWSiKMHkCeZOOwDJPgFs7SWmqHR+nKSw25vkU0XF3a954ucfEnJWP1C49HT3FzfuN22J0V6hefKqq/s6fHxfUvVzPquuoLHYmNdd7rS0Yd7XrZQ3PmBX0UVwoblTtqqCrhqIncpInhzT6Qs/aK05a9p20PUtXreWSolop3xwUj5C0taHlvDB5NAAx38V3msNN2/Zbs5vbtFV1U91fUshmk9kb3UNLiCABwaQPJzzOeKxkraKahvfS4dXDiSHR1+5qU53noo+hjvfzfS+j3cuL8i1pNX6XirfqbJqCgbU5x1RnbvZ9a+24XW22qnFVca6ClhJDRJNIGNz3ZKoWh2U6Gq9k51PLM51ydROrXVfXHAlDSd0jljPDB4+ldDdLzcb3sFhbd3unNHdW08T5fKLowMjJPPG8R5gF6+SQk0oyfPD4e4tZbT3NCDlXpRTcN+OHlY4cH2PiafininibNC9r43gOa4HIcDyIKieq9ZexKB8Gl7jZZrk2YRuZV1bWRxAHyi7Bzw7hxXbaQ3BpS1NYQB7ChwB2DcCzVRWelum0HVDZNCnUu5WTnqhP1XVfZT5WcHOeS8W9vGcpbz+qXmt6zWtaFFUVxq8M8eHDPDCb9hp60XWkulE2aG40lU9jQJX07w6MPxxxxOB516oNV6bqa11ugvtDJVNODE2oaXZ7sZVHapdVaS2U1TLPpGfSjrhcGU88bajrHOZug729gY3va8O5eNw09s52fWLTmo6rS9beHSBlUa2nnAYZcAgOycYJJIHgvStYPjl8XhYLee0dem1Ddit1Jybckkm8LC3c+OUX5LebVBcI7ZLcaZlXKAWQOkAkcDyw3meRXruuo7JZQBd7rSUe/7XrpmtJ9BVKapniq9u+l64RBplo4ZWAjiMiQ4XWbN9M2napqrUF21tLJVVMMg6undKW7oLnA4HMBoaAAOSKzSjvSfDCb9xWe01apWdrbwTm5uKy2lhJNt8OfHgaHo6+iuVOKmgq4aiJ3J8Tw5p9IX0YGcrP+zVh0dtkuWirBVy1NnlY4uG+XNiIYHg92QSW57e1X/wAQDx7Fb1qXopJJ5TWTN6PqT1OhKpOO7KLcWk8rK7H2HU1Gq9N0k8lLVX2hiliO69j6hoc09xBK+m3Xq1XbfNsuNNV9Xjf6mUP3c8s48yzZe7XT3Ha5f4XaNOozvucaZk3VFvFv2QnB5cvSrc2Q2WmtVPcnxaCdph75GMLHVHW9e0NyHZwMYLiF9qttCnT3s8cLs+OTF6Zr1xf3joOCUFJrP0s8P9O7x8SfVldR0FO6qraqOCFvtnyPDWj0lfJatSWG9h/1Iu9HWdX7YQzNcR5wFS+1wyal2r2LRF4rn09lljjkcwHAe5znZ9J3Q0E8lNLNsj0TpvUzL3ZpZqapip3BtK2pJaQcgvI9sQRwxy9K8+ghGCc28tZXDgfaOsXVxd1KdvTi4Qe623h5xltLs7O0nNDerVdHSx26401S6AgSCKQOLM9+OXIr1R6jsU3snqrvRv8AYYJqN2Zp6oDmXceHpVNdHh0Ed+1iG7gJmiOBgHG9Io5pZkD6XanLExhZ7GmyWgYOTL+pfT5GlJxzyx7SyW09SVvRrbizPfys8txN+3BoKPV+lZHiOPUNuc53ICqYc/GvurLlQUFN7MrqyCnhHHrJJA1vrKyHRR7PBomaOttVY/Usr3tp5mNd1Wd7yO3dIAxkc1PL7s615c9n+loqilFZLa3zSzUEk/lmNzgY24PBxDQQRnIzwXudlCLScscccSzttrbu4pzlCgpNRUlutvGWlh8OfXwyXtbNS6fvLiy03ikq3fcxTNcfVnK9t1v1mskYlu10paNruRmlDM+bKpDQ110Ta9c0MF12dTaZvcjDFSuDndU9zuGd3hgniAePnXw6OsVt2obSNR1WuJZZnUUjm09G6UtG6HubgDngBo4DvXh2kYybk2klnq9hdw2lrVaUIUoxdWcnFLLSWFl5yk0+7BoK33K33anFVbq2GqiPJ8Uge0+kL6sYOVn/AEfBHoXbhJpHTNRJLa6th6+LfLmx/Yy/J7MtcAM88OwtAnvVvXpKlJYeU1lGc0fUpalSk6kd2cJOLSeVldj7Do9aaej1Vpa56flkfGK6nfEHsOC0kcD68L8+6+jnttdU2+pG7NTSvhkHc5pIPxhfo+4ZasHbaKeGm2o6iggYGMFXkAd5aCfjJUK9LVjB0KF6vrJuL701leWCd+iW9mrivZfytKXg08Pzz7CFoiKCycwiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiA45hSvZlrms2fauo77TuJh3urqo88JIXHyvSOY8Qoqiu7O7q2NeFzReJRaafgWl7Z0r+3nbV1mMk014n6N2y4Ut1t9PcqGZstPUxtkje08HNIyCvsOCOCzp0XNpbqulfoG7VOZaYdZQOefbR/bMHmPEeBWiSD39i6y2f1mlrthTvKfWuK7GuaOTtf0erod/Us6vU+D7U+T/AO9ZnDpE6Dba7nHrK3Q7tPXuEVWGjg2XHBx/CAx5wqY32fdN9a3Td6IXC21FJuROdJGQ3rYhIwO+1JaeBwcHCzXZL9tEuuuhoeSz2KKpjqCyoJs8W6yJp8qT2vIt4jvyFkdR2yo6C6NvcUpTdR7qcVnj1J5xgjqr0Uz2luK97Z1401Fb0lL2tY6v1Jb0c9Amnp5Nc3KHElS0xUTXYOI8+U/0kYHgFemMjjyXqpoI6aFkEMTI2MAa1jGgNaByAA5Be455q/rVnXm5szOj6VS0e0ha0uOOb7X1srTVexGy3+9P1BbLtXWWsqMmd1I4BspPMkdhPbjmvv0xsh0zpyz19olfU3Bl0bu1bqmTJf5gODTnjkcc44qednNO1Vdeo47ueBSOiWEKrrqkt55z6+fDlxKed0crJ1phj1Pd2W50m+aMSjdPHlnt8/NSi+7KdP3TSFNoyjfJb6GllbKzqsOcSMkkl3MkkklTkApxxxKO4qtpuXI8U9A06lGcYUklJYfPl2c+C7kVZZthdNZblR3CLV15eKOVsjYXTfY3Y+1I5YXhUbB6SS9V96o9V3Simr5nzSex3BntnF27kcxkq1jy7Qnfjmq/Kauc5PH/AI9p24oej4J5XF88Y7SEWvZhb6ax3DT97u1deqW4FpcKx+8YyBwLD2Htz4KNR9Hq1brKGq1TeJ7XG/fbROkaGg+fHD0K2+J4lcnly5orirHOHzPpU0LT6yip0091YXF8s5w+1Z6mQet2W2us1la9XitnY+1xMhjgABYWsDgMnn9suq1PsNsd9vE1/tV2r7NWVLi+Z1I/DXuPM47M9uFZnLmE5ryripFpp9x6q6JYVoyjOmmm8vnzxjKfVw7CGaB2X2DQLZZ6F01TW1I3ZaqoIL3NzndGOAGfWpocYKceZQcua8TnKpLek8svra1o2dNUaEVGK6kVhdtiNNctR1upKfU9zoaiskL3exy1u6D9rkccKS6K0RLpA1fW6huF09lbn8bfvbm7nl58/EpVhuDwQ4XuVeco7rfAs6GkWdvW9PTjiWW85fN8+BENe7NbFr6nh+qPW09XT56iqgIEjM9nHmPBdHpPYnatPXqLUNde7ldqynGIXVEuGt7s44nHceHgrKIA8y54ZwirVFDcT4Famj2Va4VzOmnNYee9cm1ybXgVXetglnrrzU3ezX+42d9Y5z5o6Zw3SXe2x2gE9nJdtY9kNg0/pi7adoqmpc68xujqquQh0jsggY7ABk8PEqffGuBlp4qruKrW63wPlT0LTqdSVWNJJvPb188Lks9eCt3bE9Py6Kg0fPWTvbS1D6mGr3WiWNznZIHZggkL6LvsmivNstlNV6ou7au0tLIauKbdc4f3m8icYGefBWD51wnyipnOT0tEsFHdVNYwl18ly6+orqxbHqOg1FTakvmobhe6yh/i5qiA1h7Dgc8ZOPHivHVmxWyakvD9QUNzrrPXzHMslK/AecYzjsPfjmrI4AhOHYU+UVN7ezxKfsOwdF0HTTi3nrzntzzzgg+g9lNg0LPLcKaSesuE7S2SqqHZeWk5IA5DJA8eCnHDCcfMgOBxK+c5ym96Tyy+tbSjY01Rt4qMV1I9csjIonSOcAGjJJ7AsAbRr03UWur3eY8blTWPLCDkFrTugjzgArT/AEhNqlNpDTsunrXUg3m5sLAGO8qGI8HPPcSOA9fYse8SclQP0qa3SuKtPTaLzu5lLub4JepZJ66K9Eq0KdTU6yxvfRj3pPLfg3jHgERFDpMQREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREAREQBERAEREB2Wmb5VaZv9Bf6IkTUMzZW8eeDxB8CMg+db+01fqDU1kpL5bZmyQVcTZGkHlw4g+IOQfMvzu7OCvHo6bXWaVrv3H6gqt221j800r3eTBIftT3Nd8RUndHG0sNJu3ZXDxTqY4vkpdXqfIjHpI2bnq1rG+tlmpTzw63Hr9a5r1mtQOHHtXVwabtFPfajUkNGxtfUwsgkmHNzGkkD4+a7JkjJWB8Tw5pGQQeGF54XQkqVOtuykk8PK7n2o58jOdPKi2srD712MY7V8txr6W1UU1xrpWxU9Ox0kr3Hg1oGSfiX0kgFZp6Sm12CeF+z/TtWJMuzcZY3ZAA5Rgjx5+bCxG0OuUNAsZ3VV8V9Vdr6l8TMbP6JX16+haUVwf1n/hXW/h3liU3SW2UTRtfLep4HO5sfSyFw/wDq0j415v6SWyVoyNQyP81JL+lqxQuVCnzrawv/AMcPJ/Emn5qdHzn0k/NfA2wOkjskIz+6Rw/+JL+yvEdJPZKXFv7oZAB9t7Elwf8AlWKUVfnX1j7uHk/6h81OkfeT84/0m2f3yGyL75Xf5Ob9hcfvj9kf3zO/yc37CxOifOvq/wB3DyfxPPzUaT95Pzj/AEm1z0kdkg/8yOPmo5f2UHSQ2R490rv8nN+ysUInzr6x93Dyf9RX5qNH+8qecf6Ta/75LZJ98b/8pL+yuGdJPZISR+6CUY7TSS8f+VYo49yZPcnzr6x93Dyf9RX5qNI+8n5x+BtWo6SmyaGB0kd+mme0ZDG0koc7wGWgfGok3pdaaNy6h2m68UOOE++3rM/gcsf8SyuuVbV+lDXKzThuxx2Ln45b9hcUOi/RKSam5yz2vl4YS9ptODpKbJpomvkv00RcMljqSXLfA4aQvYekjskaMjUbz4Ckl/ZWJ+PcnHuVwulfWUsbkPJ/1Fu+inSG8+kn5x+Bsqj6UGy+oqJYJquspmRnDZZKZxbJ4jdyfWAvvHSQ2R490zh/8SX9lYnReYdK2sxWHCD9T+JWfRVo7eVOa9a/pNrjpI7JD/5lePPRy/srn98hskHLUrz/APEl/ZWJ0Xr519X+7h5P4lPmo0f7yp5x/pNsfvkNkn3yu/yc37Kfvkdkn3yu/wAnL+ysTonzr6x93DyfxKfNRpP3k/OP9Jtk9JHZH98j/wDKTfsL1R9JTZK8uBv0rADgF1JL5Xm8lYrXHoVfnX1j7uHk/wCoLop0f7yfmv6TakvST2TxRF8d/llcBkMbSS5PgMtAVa636WFTVQvo9EWl9MXDHsqqILh5mDIB8ST5lnVPSrC/6StbvqbpRcYZ64pp+bbx6i+sOjTRLKoqs1Kpjqk+Hkks+vJ9NzulwvFbLcrnVS1NVO7efJI7JcV86LjmFoVSrKrJzm8t9Zv9KnClBQgsJHKIi+R9AiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiImcFGsli6P297Q9F0bbfR11PXUrBiOOtjdJujuDg4OA8MqS/vstpH9mWH/Lzf8A+qpTATj3rZLba3WrSmqVK5korks5x5mt3OyGiXlR1qtvFyfN8s+RaOpekftL1LQPtzqult0Ug3Xmijcx5H4TnOI9GFV73vkeZHuLnOOSScklEWO1DVr3VZqd5Vc2uWXnHh2GS07SLLSYOFlSUE+eFz8e0IiLGGTCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiAIiIAiIgCIiA/9k="

def get_ou_criar_saldo(pessoa_id):
    s = SaldoAluno.query.filter_by(pessoa_id=pessoa_id).first()
    if not s:
        s = SaldoAluno(pessoa_id=pessoa_id, saldo=0.0)
        db.session.add(s)
        db.session.commit()
    return s

def buscar_pessoa_por_matricula(matricula):
    mat = matricula.strip().upper()
    return Pessoa.query.filter(db.func.upper(Pessoa.matricula) == mat).first()

@app.route("/aluno/<matricula>")
def pagina_aluno(matricula):
    p = buscar_pessoa_por_matricula(matricula)
    if not p:
        return "<h2 style=\'font-family:sans-serif;color:#c00;padding:40px\'>Matrícula não encontrada.</h2>", 404
    s = get_ou_criar_saldo(p.id)
    return render_template("aluno.html", pessoa=p, saldo=s,
                           pix_chave=PIX_CHAVE, pix_nome=PIX_NOME,
                           matricula=matricula, logo_b64=LOGO_B64_GLOBAL)

@app.route("/tablet")
def pagina_tablet():
    return render_template("tablet.html")

@app.route("/api/saldo/aluno/<matricula>")
def api_aluno_por_matricula(matricula):
    p = buscar_pessoa_por_matricula(matricula)
    if not p:
        return jsonify({"erro": "Matrícula não encontrada"}), 404
    s = get_ou_criar_saldo(p.id)
    recargas = RecargaSaldo.query.filter_by(pessoa_id=p.id, confirmado=True)\
                                 .order_by(RecargaSaldo.data.desc()).limit(10).all()
    vendas   = Venda.query.filter_by(pessoa_id=p.id)\
                          .order_by(Venda.data.desc()).limit(10).all()
    pedidos  = PedidoTablet.query.filter_by(pessoa_id=p.id)\
                                 .order_by(PedidoTablet.data.desc()).limit(10).all()
    historico = []
    for r in recargas:
        historico.append({"tipo":"recarga","data":r.data.strftime("%d/%m/%Y %H:%M"),
                          "valor":r.valor,"obs":r.obs or ""})
    for v in vendas:
        historico.append({"tipo":"compra","data":v.data.strftime("%d/%m/%Y %H:%M"),
                          "valor":-v.total,
                          "obs":", ".join(f"{i.produto_nome} x{i.qtd}" for i in v.itens)})
    for pd in pedidos:
        itens_p = json.loads(pd.itens_json or "[]")
        historico.append({"tipo":"pedido_tablet","data":pd.data.strftime("%d/%m/%Y %H:%M"),
                          "valor":-pd.total,
                          "obs":", ".join(f"{i[\'nome\']} x{i[\'qtd\']}" for i in itens_p)})
    historico.sort(key=lambda x: x["data"], reverse=True)
    return jsonify({"pessoa_id":p.id,"nome":p.nome,"turma":p.turma_cargo or "",
                    "matricula":p.matricula or "","saldo":s.saldo,
                    "historico":historico[:20]})

@app.route("/api/saldo/<int:pid>/info")
def api_saldo_info(pid):
    p = Pessoa.query.get(pid)
    if not p: return jsonify({"erro":"Não encontrado"}), 404
    s = get_ou_criar_saldo(pid)
    url = (request.host_url.rstrip("/") + f"/aluno/{p.matricula}") if p.matricula else None
    return jsonify({"saldo":s.saldo,"matricula":p.matricula or "","url":url})

@app.route("/api/saldo/recargas/pendentes")
def recargas_pendentes():
    recargas = RecargaSaldo.query.filter_by(confirmado=False)\
                                 .order_by(RecargaSaldo.data.desc()).all()
    return jsonify([{"id":r.id,"pessoa_id":r.pessoa_id,"nome":r.pessoa.nome,
                     "turma":r.pessoa.turma_cargo or "","valor":r.valor,
                     "data":r.data.strftime("%d/%m/%Y %H:%M"),"obs":r.obs or ""}
                    for r in recargas])

@app.route("/api/saldo/recarga/solicitar", methods=["POST"])
def solicitar_recarga():
    d = request.json
    valor = float(d.get("valor", 0))
    if valor <= 0: return jsonify({"erro":"Valor inválido"}), 400
    r = RecargaSaldo(pessoa_id=d.get("pessoa_id"), valor=valor,
                     confirmado=False, obs=d.get("obs",""))
    db.session.add(r)
    db.session.commit()
    return jsonify({"ok":True,"id":r.id})

@app.route("/api/saldo/recarga/<int:rid>/confirmar", methods=["POST"])
def confirmar_recarga(rid):
    r = RecargaSaldo.query.get(rid)
    if not r: return jsonify({"erro":"Não encontrado"}), 404
    if r.confirmado: return jsonify({"erro":"Já confirmado"}), 400
    r.confirmado = True
    s = get_ou_criar_saldo(r.pessoa_id)
    s.saldo += r.valor
    db.session.commit()
    return jsonify({"ok":True,"novo_saldo":s.saldo})

@app.route("/api/saldo/recarga/<int:rid>/rejeitar", methods=["POST"])
def rejeitar_recarga(rid):
    r = RecargaSaldo.query.get(rid)
    if not r: return jsonify({"erro":"Não encontrado"}), 404
    db.session.delete(r)
    db.session.commit()
    return jsonify({"ok":True})

@app.route("/api/saldo/<int:pid>/adicionar", methods=["POST"])
def adicionar_saldo(pid):
    valor = float(request.json.get("valor",0))
    obs   = request.json.get("obs","Adicionado manualmente")
    if valor <= 0: return jsonify({"erro":"Valor inválido"}), 400
    s = get_ou_criar_saldo(pid)
    s.saldo += valor
    r = RecargaSaldo(pessoa_id=pid, valor=valor, confirmado=True, obs=obs)
    db.session.add(r)
    db.session.commit()
    return jsonify({"ok":True,"novo_saldo":s.saldo})

@app.route("/api/tablet/pedido", methods=["POST"])
def pedido_tablet():
    d = request.json
    matricula = d.get("matricula","").strip()
    carrinho  = d.get("carrinho",[])
    if not matricula or not carrinho:
        return jsonify({"erro":"Dados incompletos"}), 400
    p = buscar_pessoa_por_matricula(matricula)
    if not p: return jsonify({"erro":"Matrícula não encontrada"}), 404
    s = get_ou_criar_saldo(p.id)
    total = 0.0
    itens_ok = []
    for item in carrinho:
        prod = Produto.query.get(item["id"])
        if not prod or prod.estoque < item["qtd"]:
            return jsonify({"erro":f"Estoque insuficiente: {item[\'nome\']}"}), 400
        sub = prod.preco * item["qtd"]
        total += sub
        itens_ok.append({"id":prod.id,"nome":prod.nome,"preco":prod.preco,
                          "qtd":item["qtd"],"sub":sub})
    if s.saldo < total:
        return jsonify({"erro":f"Saldo insuficiente. Disponível: R$ {s.saldo:.2f}"}), 400
    for item in itens_ok:
        Produto.query.get(item["id"]).estoque -= item["qtd"]
    s.saldo -= total
    pedido = PedidoTablet(pessoa_id=p.id, total=total, status="aguardando",
                          itens_json=json.dumps(itens_ok, ensure_ascii=False))
    db.session.add(pedido)
    venda = Venda(total=total, forma_pagamento="Saldo", pago=True, pessoa_id=p.id)
    db.session.add(venda)
    db.session.flush()
    for item in itens_ok:
        db.session.add(ItemVenda(venda_id=venda.id, produto_nome=item["nome"],
                                  qtd=item["qtd"], preco_unit=item["preco"],
                                  subtotal=item["sub"]))
    db.session.commit()
    return jsonify({"ok":True,"total":total,"novo_saldo":s.saldo,
                    "pedido_id":pedido.id,"nome":p.nome,"itens":itens_ok})

@app.route("/api/tablet/pedidos")
def listar_pedidos_tablet():
    status = request.args.get("status","aguardando")
    pedidos = PedidoTablet.query.filter_by(status=status)\
                                 .order_by(PedidoTablet.data.desc()).limit(50).all()
    return jsonify([{"id":p.id,"nome":p.pessoa.nome,"turma":p.pessoa.turma_cargo or "",
                     "data":p.data.strftime("%H:%M"),"total":p.total,"status":p.status,
                     "itens":json.loads(p.itens_json or "[]")} for p in pedidos])

@app.route("/api/tablet/pedidos/<int:pid>/entregar", methods=["POST"])
def entregar_pedido(pid):
    p = PedidoTablet.query.get(pid)
    if not p: return jsonify({"erro":"Não encontrado"}), 404
    p.status = "entregue"
    db.session.commit()
    return jsonify({"ok":True})


if __name__ == "__main__":
    app.run(debug=True)
